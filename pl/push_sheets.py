#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ローカルの .xlsx を、Driveの【ネイティブのスプレッドシート】へ書き写す

    GOOGLE_SA_KEY_FILE=/tmp/sa.json python3 push_sheets.py 22期
    GOOGLE_SA_KEY_FILE=/tmp/sa.json python3 push_sheets.py 22期 --id <ファイルID>
    GOOGLE_SA_KEY_FILE=/tmp/sa.json python3 push_sheets.py 22期 --dry-run

--- なぜ push.py と別なのか ----------------------------------------------
push.py は Drive の files.update に .xlsx を丸ごと投げる。相手が .xlsx なら
これでいいが、【ネイティブのスプレッドシートには使えない】。
2026-08-21 に試したら 500 Internal Error（native_probe.py に記録）。

一方 .xlsx のまま Drive に置くと、ブラウザで開くたびに Google がその場で
スプレッドシートへ変換する（DOCS_EVERYWHERE_IMPORT）。13シート3万セルだと
これが重く、同日 QUOTA_EXCEEDED で開けなくなった。

なのでネイティブ化したうえで、こちらは Sheets API でセルを書く。
Sheets API なら値・数式・数値書式・色・タブ追加が全部通る（実測済み）。

--- 作り -----------------------------------------------------------------
fill2.py / build22.py が作る .xlsx を openpyxl で読んで、そのままシートへ写す。
転記のロジックには一切触らない。書き換えたのは「Driveへ送る部分」だけ。

    ① タブを合わせる（足りなければ追加、余っていれば削除）
    ② 値と数式を書く   values().batchUpdate + USER_ENTERED
       USER_ENTERED にすると "=SUM(...)" が数式として解釈される
    ③ 書式を書く       spreadsheets().batchUpdate
       数値書式（赤字マイナス）・太字・文字色・背景色・罫線・
       列幅・結合セル・ウィンドウ枠固定

--- 注意 -----------------------------------------------------------------
・丸ごと差し替えなので、Drive側で手入力した内容は消える（push.py と同じ）
・1回のリクエストが大きくなりすぎないよう、書式は塊にまとめて送っている
"""
import os
import sys

import openpyxl
from openpyxl.utils import get_column_letter

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# 期: (ローカルの.xlsx, ネイティブのスプレッドシートのID)
# ★push.py の TARGETS とは別物。あちらは .xlsx のファイルID。
# どちらもDriveの【ネイティブのスプレッドシート】。2026年損益計算書フォルダにある。
#   21期  https://docs.google.com/spreadsheets/d/1dODfqd_EV4iZGSyepi1HZBTFDbA36YxnR2TxwBIXfRk/
#   22期  https://docs.google.com/spreadsheets/d/1sq6pulq1dT3AWo2zzQTRukajlUar_QHULOOqjjOLL24/
# 21期には .xlsx 版（push.py の SHEET_ID_21）も残してある。当面は両方に反映する。
TARGETS = {
    "21期": ("損益計算書_21期テスト版.xlsx",
            os.environ.get("GSHEET_ID_21", "1dODfqd_EV4iZGSyepi1HZBTFDbA36YxnR2TxwBIXfRk")),
    "22期": ("損益計算書_22期.xlsx",
            os.environ.get("GSHEET_ID_22", "1sq6pulq1dT3AWo2zzQTRukajlUar_QHULOOqjjOLL24")),
}


def _creds():
    from google.oauth2 import service_account
    path = os.environ.get("GOOGLE_SA_KEY_FILE")
    if not path or not os.path.exists(path):
        sys.exit("GOOGLE_SA_KEY_FILE にサービスアカウントのJSONを指定してください。"
                 "取り方は push.py の冒頭を参照")
    return service_account.Credentials.from_service_account_file(path, scopes=SCOPES)


def _rgb(color):
    """openpyxl の色 → Sheets API の {red,green,blue}。取れなければ None。"""
    if color is None or color.type != "rgb" or not color.rgb:
        return None
    s = str(color.rgb)
    if len(s) == 8:                       # AARRGGBB
        s = s[2:]
    if len(s) != 6:
        return None
    return {"red": int(s[0:2], 16) / 255,
            "green": int(s[2:4], 16) / 255,
            "blue": int(s[4:6], 16) / 255}


def _fmt(cell):
    """1セルぶんの userEnteredFormat。既定のままなら None。"""
    f = {}
    if cell.number_format and cell.number_format != "General":
        n = cell.number_format
        kind = "PERCENT" if "%" in n else "NUMBER"
        f["numberFormat"] = {"type": kind, "pattern": n}
    bg = _rgb(cell.fill.fgColor) if cell.fill and cell.fill.patternType else None
    if bg:
        f["backgroundColor"] = bg
    t = {}
    if cell.font:
        if cell.font.bold:
            t["bold"] = True
        if cell.font.italic:
            t["italic"] = True
        if cell.font.size and float(cell.font.size) != 11:
            t["fontSize"] = int(float(cell.font.size))
        fg = _rgb(cell.font.color)
        if fg:
            t["foregroundColor"] = fg
    if t:
        f["textFormat"] = t
    if cell.alignment and cell.alignment.horizontal:
        f["horizontalAlignment"] = cell.alignment.horizontal.upper()
    if cell.border and cell.border.left and cell.border.left.style:
        line = {"style": "SOLID", "width": 1,
                "color": _rgb(cell.border.left.color) or {"red": .7, "green": .78, "blue": .91}}
        f["borders"] = {k: line for k in ("top", "bottom", "left", "right")}
    return f or None


def _sync_tabs(svc, sid, names):
    """タブ名をローカルに合わせる。→ {タブ名: sheetId}"""
    meta = svc.spreadsheets().get(spreadsheetId=sid,
                                  fields="sheets.properties").execute()
    have = {s["properties"]["title"]: s["properties"]["sheetId"]
            for s in meta["sheets"]}
    req = []
    for n in names:
        if n not in have:
            req.append({"addSheet": {"properties": {"title": n}}})
    # 余分なタブは消す。ただし全部消すとエラーになるので、追加のあとに消す
    extra = [t for t in have if t not in names]
    if req:
        svc.spreadsheets().batchUpdate(spreadsheetId=sid, body={"requests": req}).execute()
        meta = svc.spreadsheets().get(spreadsheetId=sid,
                                      fields="sheets.properties").execute()
        have = {s["properties"]["title"]: s["properties"]["sheetId"]
                for s in meta["sheets"]}
    if extra:
        svc.spreadsheets().batchUpdate(spreadsheetId=sid, body={"requests": [
            {"deleteSheet": {"sheetId": have[t]}} for t in extra]}).execute()
        for t in extra:
            have.pop(t)
    # 並び順もローカルに合わせる
    order = [{"updateSheetProperties": {
        "properties": {"sheetId": have[n], "index": i},
        "fields": "index"}} for i, n in enumerate(names)]
    svc.spreadsheets().batchUpdate(spreadsheetId=sid, body={"requests": order}).execute()
    return have


def _split_values(ws):
    """シートを (数式だけ, 数式以外) の2枚に分ける。

    ★なぜ分けるのか（2026-08-21 に踏んだ）
      全部を USER_ENTERED で送ると、数式は効くが、明細ログの
      '2026-06-30' のような【日付に見える文字列】をGoogleが日付として
      読み替えてしまい、46203 というシリアル値になる。2,207セルが化けた。
      RAW で送れば文字列のまま入るが、今度は数式が効かない。
      なので数式（= で始まるもの）だけ USER_ENTERED、残りは RAW で送る。
    """
    formula, plain = [], []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        pr = []
        run, start = [], None                 # 横に続く数式はまとめて1範囲にする
        for c in row:
            v = "" if c.value is None else c.value
            is_f = isinstance(v, str) and v.startswith("=")
            pr.append("" if is_f else v)
            if is_f:
                if start is None:
                    start = c.column
                run.append(v)
            elif run:
                formula.append((ws.title, c.row, start, run))
                run, start = [], None
        if run:
            formula.append((ws.title, row[0].row, start, run))
        plain.append(pr)
    return formula, plain


def _format_requests(ws, sheet_id):
    """書式のリクエスト。同じ書式が続くセルは1本にまとめて送る。"""
    req = [
        # まず全面をまっさらに戻す（前回の残りを消す）
        {"repeatCell": {"range": {"sheetId": sheet_id},
                        "cell": {"userEnteredFormat": {}},
                        "fields": "userEnteredFormat"}},
        {"unmergeCells": {"range": {"sheetId": sheet_id}}},
    ]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        run_fmt, run_start = None, None

        def flush(end):
            if run_fmt and run_start is not None:
                req.append({"repeatCell": {
                    "range": {"sheetId": sheet_id,
                              "startRowIndex": row[0].row - 1, "endRowIndex": row[0].row,
                              "startColumnIndex": run_start, "endColumnIndex": end},
                    "cell": {"userEnteredFormat": run_fmt},
                    "fields": "userEnteredFormat"}})

        for c in row:
            f = _fmt(c)
            if f != run_fmt:
                flush(c.column - 1)
                run_fmt, run_start = f, c.column - 1
        flush(ws.max_column)

    # 列幅
    for letter, dim in ws.column_dimensions.items():
        if not dim.width:
            continue
        try:
            i = openpyxl.utils.column_index_from_string(letter) - 1
        except ValueError:
            continue
        req.append({"updateDimensionProperties": {
            "range": {"sheetId": sheet_id, "dimension": "COLUMNS",
                      "startIndex": i, "endIndex": i + 1},
            "properties": {"pixelSize": int(dim.width * 7.5)},
            "fields": "pixelSize"}})

    # 結合セル
    for m in ws.merged_cells.ranges:
        req.append({"mergeCells": {
            "range": {"sheetId": sheet_id,
                      "startRowIndex": m.min_row - 1, "endRowIndex": m.max_row,
                      "startColumnIndex": m.min_col - 1, "endColumnIndex": m.max_col},
            "mergeType": "MERGE_ALL"}})

    # ウィンドウ枠固定
    if ws.freeze_panes:
        cell = ws[ws.freeze_panes]
        req.append({"updateSheetProperties": {
            "properties": {"sheetId": sheet_id, "gridProperties": {
                "frozenRowCount": cell.row - 1,
                "frozenColumnCount": cell.column - 1}},
            "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount"}})
    return req


def _chunks(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


def push(period, sid=None, dry=False):
    local, default_id = TARGETS[period]
    sid = sid or default_id
    if not sid:
        sys.exit(f"{period} のスプレッドシートIDが未設定です。"
                 f"--id で渡すか、環境変数 GSHEET_ID_* を設定してください")
    if not os.path.exists(local):
        sys.exit(f"{local} がありません。先に fill2.py / build22.py を実行してください")

    wb = openpyxl.load_workbook(local)
    names = wb.sheetnames
    cells = sum(w.max_row * w.max_column for w in wb.worksheets)
    print(f"元: {local}  タブ{len(names)}枚 / 約{cells:,}セル")
    if dry:
        print("--dry-run なのでここまで。書き込みません。")
        print("  タブ:", names)
        return

    from googleapiclient.discovery import build
    svc = build("sheets", "v4", credentials=_creds(), cache_discovery=False)
    meta = svc.spreadsheets().get(spreadsheetId=sid,
                                  fields="properties.title").execute()
    print(f"先: {meta['properties']['title']}  ({sid})")

    ids = _sync_tabs(svc, sid, names)
    print(f"  タブを合わせました（{len(ids)}枚）")

    # ★書き込む前に、これから書く範囲より下と右を消す（2026-08-22 追加）。
    #   行を減らしたときに、前回の中身が下に居残る。実際に売上原価の行を
    #   81行削ったら、りゅうちゃんの【指標】以下が二重に残った（748セル）。
    #   values().clear は「値だけ」消す。書式は次の書式反映で上書きされる。
    clear = []
    for n_ in names:
        ws = wb[n_]
        clear.append(f"'{n_}'!A{ws.max_row + 1}:ZZ")
        clear.append(f"'{n_}'!{get_column_letter(ws.max_column + 1)}1:ZZ")
    svc.spreadsheets().values().batchClear(
        spreadsheetId=sid, body={"ranges": clear}).execute()
    print(f"  前回の残りを消しました（{len(clear)}範囲）")

    # ② 値と数式。★2回に分ける（_split_values のコメント参照）
    #   先に RAW で全部書いてから、数式のセルだけ USER_ENTERED で上書きする。
    #   順番が逆だと、RAW が数式を "=SUM(...)" という文字列に戻してしまう。
    fx, pl = [], {}
    for n in names:
        f, p = _split_values(wb[n])
        fx += f
        pl[n] = p
    svc.spreadsheets().values().batchUpdate(spreadsheetId=sid, body={
        "valueInputOption": "RAW",
        "data": [{"range": f"'{n}'!A1", "values": pl[n]} for n in names]}).execute()
    n_fx = sum(len(run) for _, _, _, run in fx)
    for part in _chunks(fx, 2000):
        svc.spreadsheets().values().batchUpdate(spreadsheetId=sid, body={
            "valueInputOption": "USER_ENTERED",
            "data": [{"range": f"'{t}'!{get_column_letter(col)}{row}",
                      "values": [run]} for t, row, col, run in part]}).execute()
    print(f"  値を書きました（{sum(len(v) for v in pl.values()):,}行"
          f"／うち数式 {n_fx:,}セル・{len(fx):,}範囲）")

    # ③ 書式
    total = 0
    for n in names:
        req = _format_requests(wb[n], ids[n])
        for part in _chunks(req, 500):      # 1回が大きくなりすぎないように割る
            svc.spreadsheets().batchUpdate(spreadsheetId=sid,
                                           body={"requests": part}).execute()
        total += len(req)
    print(f"  書式を書きました（{total:,}件）")
    print(f"\nhttps://docs.google.com/spreadsheets/d/{sid}/edit")


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    sid = None
    if "--id" in sys.argv:
        sid = sys.argv[sys.argv.index("--id") + 1]
        args = [a for a in args if a != sid]
    if len(args) != 1 or args[0] not in TARGETS:
        sys.exit(f"使い方: python3 push_sheets.py [{' | '.join(TARGETS)}] "
                 f"[--id <ファイルID>] [--dry-run]")
    push(args[0], sid, "--dry-run" in sys.argv)
