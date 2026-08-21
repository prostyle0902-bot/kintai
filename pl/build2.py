#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""損益計算書 21期テスト版 v2 — 共通行＋既存スプシ由来の取引先別行"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sales

TABS = ["りゅうちゃん", "もも焼きJAPAN", "韓国酒場ハナ", "さわら十三里屋",
        "タコとハイボール", "焼きたて屋", "神栖横丁", "鳥害対策課", "業務課", "本部"]
# 共通行を使わないタブ（既存スプシの構成がまったく違うため専用行のみ）
NO_COMMON = {"本部"}
MONTHS = ["9月","10月","11月","12月","1月","2月","3月","4月","5月","6月","7月","8月"]
MCOL = {m: get_column_letter(3 + i) for i, m in enumerate(MONTHS)}

SEC, IN, FM = "SEC", "IN", "FM"

# 表示書式。マイナスは赤字（利用者指示 2026-08-20）
NUMFMT = "#,##0;[Red]-#,##0"
PCTFMT = "0.0%;[Red]-0.0%"

# 期の表記。new_wb(period=...) で切り替える
PERIODS = {"21期": "21期（2025.9〜2026.8）", "22期": "22期（2026.9〜2027.8）"}
PERIOD = "21期"

# ---- 共通レイアウト（仕様【1】） ----
COMMON_COGS = ["仕入（鹿島食品8%）","仕入（鹿島食品10%）","仕入（藤原ストア）","仕入（平洋酒店）",
               "仕入（六角堂）","仕入（freeeカード）","仕入（現金・米）","仕入（ネット食材）",
               "仕入（やまなか）","仕入（その他）"]
COMMON_SGA = ["人件費（店長）","人件費（アルバイト）","法定福利費","旅費・交通費","接待交際費",
              "広告宣伝費（共通宣伝費）","地代家賃（賃料）","地代家賃（駐車場利用料）","地代家賃（共益費）",
              "水道光熱費（電気料金）","水道光熱費（水道料金）","水道光熱費（ガス料金）","通信費",
              "消耗品費（freeeカード）","消耗品費","租税公課","減価償却費","保険料",
              "支払手数料（AirPay）","支払手数料（AirPayQR）","警備費","雑費","その他経費"]

# ---- 既存スプシ由来の店舗別 追加行 ----
EXTRA_COGS = {
 # 3件とも既存21期PLにあるのを写した（2026-08-21）
 "りゅうちゃん": ["仕入（鹿島食品）","仕入（藤原ストアー値引き）","仕入（現金）"],
 # 仕入（新場水産）は 2026-08-21 新設（利用者指示「その他の下に」＝仕入（その他）の下）
 "もも焼きJAPAN": ["仕入（新場水産）",
              "仕入（鹿島食品）","仕入（山食)","仕入（鈴喜）","仕入（日本食研）",
              "仕入（東京さえき）","仕入（京兼醸造）","仕入（東総食肉）","仕入（現金）"],
 "韓国酒場ハナ": ["仕入（鹿島食品）","仕入（日本食研）","仕入（山中ストアー）","仕入（東総食肉）",
             "仕入（CQREE）","仕入（現金）"],
 "さわら十三里屋": ["仕入（なめファ8％）","仕入（なめファ10％）","仕入（鹿島食品）"],
 # 仕入（フラバーダイニング）は 2026-08-21 新設（利用者指示「その他の下に」）
 "タコとハイボール": ["仕入（フラバーダイニング）",
                "仕入（インフォマート8％）","仕入（インフォマート10％）",
                "仕入（インフォマート利用料）","仕入（鹿島食品）","仕入（日本食研）","仕入（現金）"],
 "焼きたて屋": ["仕入（インフォマート8％）","仕入（インフォマート10％）","仕入（鹿島食品）","仕入（現金）"],
 "神栖横丁": ["仕入（ビーエム）","ALSOK","ウゴーク","クリーズコネクト（ゴミ）","ダスキン",
          "グリスト清掃業務課","銚子杉野","千葉銀リース","三井住友リース","鹿島食品",
          "東総食肉センター","EMBER","藤本商会本店",
          # アントラーズビアガーデン2026の飲食代。取引先名で行を新設（利用者指示 2026-08-20）
          "ガキゲン","居酒屋ドリーム","鉄板焼次郎","ツマギアンズ",
          # 以下5件は既存21期PLにあるのを写した（2026-08-21）。既存の並び順に合わせている
          "鹿島アントラーズ","シオン","ALホーム","セキネネオン","タカギ",
          "その他、仕入（小口）"],
 "鳥害対策課": ["仕入（日本鳩対策センター）","外注費（西尾レントオール）","外注費（プロエンジニア）",
           "外注費（ロープワーク）","外注費（サンクトラスト）","外注費（SUN-X）","その他、仕入（小口）"],
 "業務課": ["仕入（ビーエム）","仕入（グローバルクリーン）","仕入（ニッセーデリカ）",
         "仕入（江東微生物研究所）","仕入（ケミカルテクノロジー）","仕入（モノタロウ）",
         "外注費（ロープワーク）","外注費（増田）","外注費（OnebyOne）","外注費（SIBS）",
         "外注費（ライズ）","外注費（ミノアカ）","外注費（SUN-X）","外注費（ピカピカ）",
         "西尾レントオール","総合環境分析","ササモト技研","トヨタファイナンス（燃料）",
         "陸自総合（高速代）","その他、仕入（小口）"],
 # 本部は既存スプシ「本部 損益計算書2025.9~2026.8」の行構成そのまま。
 # セキショウホンダ・ウェイブは既存に行がないため取引先名で新設（利用者指示 2026-08-18）
 "本部": ["JCBカード","三井住友カード","JBQ","TMAP","オニオン新聞","ジョイント","スギヤマ","セコム",
        "プロセミ","ヤマト運輸","椎名環境","テイクアクション","ニッサン","高塚誠","室田石油","SCPN",
        "バイオラボ","千葉銀リース","ジブラルタ生命","NS会議所（上乗せ労災）","帝国データバンク",
        "リコーリース","三井住友リース","徳栄自動車","ネットバンク手数料","五十嵐商会","いいじま",
        "鹿島アントラーズ","青野税理士","商工会費","フォーゲル弁護士事務所","ピープルフォレスト",
        "リヴィン","CCMOコンサルティング","ノベルティハウス","ナビット","ゴルフフロンティア",
        "プルデンシャル生命","オフィスで野菜","クリーンシステム科学研究所","NTTファイナンス",
        "FM鹿島","チムニータウン","ロータリークラブ","ワンオーナー",
        "セキショウホンダ","ウェイブ","その他"],
}
EXTRA_SGA = {
 "りゅうちゃん": [],
 "もも焼きJAPAN": ["消耗品費（関彰商事）","その他経費（華まる支払い）"],
 "韓国酒場ハナ": ["リース料"],
 "さわら十三里屋": ["リース料（ダスキン）","廃棄物処分","通信費（USEN、Wi-Fi）","その他経費（ロイヤリティ）",
            # 既存21期PLにあるのを写した（2026-08-21）
            "タカギ（Instagramコンサル）"],
 # 千葉銀リースは 2026-08-21 新設（利用者指示）。店舗口座から毎月定額で落ちているが
 # 既存スプシに受け皿の行が無かった。神栖横丁・本部には元からある行名にそろえた。
 "タコとハイボール": ["消耗品費（関彰商事）","雑費（本部SV巡回活動費）","支払手数料（出前館）",
                "支払手数料（出前館返金）","支払手数料（UberEats）","保険衛生費",
                "千葉銀リース","ロイヤリティ"],
 "焼きたて屋": ["ごみ処分費","千葉銀リース","地代家賃（裏プレハブ賃料）","通信費（POSレジ）","通信費（POSレジ利用料）",
          "外注費（現金輸送費）","支払手数料（かめや）","支払手数料（Softbank）","支払手数料（出前館）",
          "支払手数料（出前館返金）","保険衛生費","ロイヤリティ"],
 "神栖横丁": ["通信費（USEN）","通信費（ダゾーン）","通信費（有線ネット）","通信費（郵便関係・その他）",
          "事務消耗品費","諸会費（Amazon会員）",
          # 既存21期PLにあるのを写した（2026-08-21）。共通宣伝費とは別物
          "広告宣伝費（LINE）"],
 "鳥害対策課": [],
 "本部": ["人件費　社長","人件費　純子","法定福利費","消費税","労働保険料","旅費・交通費",
        "広告宣伝費","接待交際費","地代家賃（145,000÷3）","通信費","事務消耗品費","租税公課",
        "減価償却費","水道","電気代　地域創生","電気代　東京電力","ガス代","保険料（損保ジャパン）",
        # JCB明細を取引先別に割るため新設（利用者指示 2026-08-20）
        "福利厚生費","ソフトウェア利用料",
        # 会費類の受け皿として新設（利用者指示 2026-08-20）。shokaihi.py 参照
        "諸会費",
        "雑費","その他経費"],
 "業務課": ["事務消耗品費","諸会費（日本ビルメン）","修繕費（リペアボックス）"],
}


# Labor比率の分子に使う人件費行（タブごとに行名が違う）
LABOR_ROWS = {t: ["人件費（店長）", "人件費（アルバイト）"] for t in TABS}
LABOR_ROWS["本部"] = ["人件費　社長", "人件費　純子"]


def layout_for(tab):
    """店舗ごとの行レイアウトを組み立てる"""
    if tab in NO_COMMON:
        cogs, sga = EXTRA_COGS[tab], EXTRA_SGA[tab]
    else:
        cogs = COMMON_COGS + EXTRA_COGS[tab]
        sga = COMMON_SGA + EXTRA_SGA[tab]
    L = [(SEC, "【売上】", None)]
    parts = []
    for uri, zei in sales.SALES_ROWS[tab]:
        L.append((IN, uri, None))
        if zei:
            L.append((IN, zei, None))
            parts.append("({c}{r_%s}-{c}{r_%s})" % (uri, zei))
        else:
            parts.append("{c}{r_%s}" % uri)
    # 売上合計(1) は税抜。消費税行があるものは 売上(税込)-消費税
    L.append((FM, "売上合計(1)", "=" + "+".join(parts)))
    L.append((SEC, "【売上原価】", None))
    L += [(IN, x, None) for x in cogs]
    L += [(FM, "仕入合計(2)", "=SUM({c}%s:{c}%s)" % ("{r_" + cogs[0] + "}", "{r_" + cogs[-1] + "}")),
          (SEC, "【棚卸】", None), (IN, "期首棚卸し", None), (IN, "期末棚卸し", None),
          (FM, "売上原価(a)", "={c}{r_仕入合計(2)}+{c}{r_期首棚卸し}-{c}{r_期末棚卸し}"),
          (SEC, "【売上総利益】", None),
          (FM, "売上総利益(3)", "={c}{r_売上合計(1)}-{c}{r_売上原価(a)}"),
          (SEC, "【販売費・一般管理費】", None)]
    L += [(IN, x, None) for x in sga]
    L += [(FM, "販管費合計(4)", "=SUM({c}%s:{c}%s)" % ("{r_" + sga[0] + "}", "{r_" + sga[-1] + "}")),
          (SEC, "【利益】", None),
          (FM, "営業利益(5)", "={c}{r_売上総利益(3)}-{c}{r_販管費合計(4)}"),
          (IN, "営業外収益(6)", None), (IN, "営業外費用(7)", None),
          (FM, "経常利益(8)", "={c}{r_営業利益(5)}+{c}{r_営業外収益(6)}-{c}{r_営業外費用(7)}"),
          (IN, "特別利益(9)", None), (IN, "特別損失(10)", None),
          (FM, "税引前当期利益(11)", "={c}{r_経常利益(8)}+{c}{r_特別利益(9)}-{c}{r_特別損失(10)}"),
          (IN, "法人税等(12)", None),
          (FM, "税引後当期利益", "={c}{r_税引前当期利益(11)}-{c}{r_法人税等(12)}"),
          (SEC, "【指標】", None),
          (FM, "Food比率（目標25%）", '=IFERROR({c}{r_売上原価(a)}/{c}{r_売上合計(1)},"")'),
          (FM, "Labor比率（目標35%）",
           '=IFERROR((%s)/{c}{r_売上合計(1)},"")'
           % "+".join("{c}{r_%s}" % x for x in LABOR_ROWS[tab]))]
    return L


def ridx_for(L):
    idx, r = {}, 3
    for _, label, _ in L:
        idx[label] = r; r += 1
    return idx


LAYOUTS = {t: layout_for(t) for t in TABS}
RIDX = {t: ridx_for(LAYOUTS[t]) for t in TABS}

F_SEC = PatternFill("solid", fgColor="1F3864"); F_IN = PatternFill("solid", fgColor="FFF7D6")
F_FM = PatternFill("solid", fgColor="E8EDF5"); F_HDR = PatternFill("solid", fgColor="305496")
F_YR = PatternFill("solid", fgColor="D6E4F0"); F_EXTRA = PatternFill("solid", fgColor="FFE9D6")
THIN = Side(style="thin", color="B4C6E7")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def render(tpl, col, ridx):
    s = tpl.replace("{c}", col)
    for label, r in sorted(ridx.items(), key=lambda kv: -len(kv[0])):
        s = s.replace("{r_%s}" % label, str(r))
    return s


def build_pl_tab(ws, tab):
    L, ridx = LAYOUTS[tab], RIDX[tab]
    extra = set() if tab in NO_COMMON else set(EXTRA_COGS[tab]) | set(EXTRA_SGA[tab])
    ws.cell(1, 1, f"{tab}　損益計算書　{PERIODS[PERIOD]}／単位：円・税抜（円未満切り捨て）"
            ).font = Font(bold=True, size=13)
    for j, h in enumerate(["勘定科目", "年計"] + MONTHS, start=1):
        c = ws.cell(2, j, h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = F_HDR; c.alignment = Alignment(horizontal="center"); c.border = BORD
    for kind, label, tpl in L:
        r = ridx[label]
        a = ws.cell(r, 1, label); a.border = BORD
        if kind == SEC:
            a.font = Font(bold=True, color="FFFFFF"); a.fill = F_SEC
            for j in range(2, 15):
                cc = ws.cell(r, j); cc.fill = F_SEC; cc.border = BORD
            continue
        a.font = Font(bold=(kind == FM))
        if label in extra:
            a.fill = F_EXTRA          # 既存スプシ由来の店舗固有行
        ratio = "比率" in label
        b = ws.cell(r, 2)
        b.value = f'=IFERROR(AVERAGE(C{r}:N{r}),"")' if ratio else f"=SUM(C{r}:N{r})"
        b.fill = F_YR; b.font = Font(bold=True); b.border = BORD
        b.number_format = PCTFMT if ratio else NUMFMT
        for m in MONTHS:
            col = MCOL[m]; cc = ws[f"{col}{r}"]
            cc.border = BORD; cc.number_format = PCTFMT if ratio else NUMFMT
            if kind == FM:
                cc.value = render(tpl, col, ridx); cc.fill = F_FM; cc.font = Font(bold=True)
            else:
                cc.fill = F_IN
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 12
    for m in MONTHS:
        ws.column_dimensions[MCOL[m]].width = 11
    ws.freeze_panes = "C3"


def build_summary(ws):
    ws.cell(1, 1, f"全体サマリー　{PERIODS[PERIOD]}／単位：円・税抜").font = Font(bold=True, size=14)
    ws.cell(2, 1, "各店舗タブを参照しています（手入力しないでください）").font = Font(italic=True, color="808080")
    keys = ["売上合計(1)","売上原価(a)","売上総利益(3)","販管費合計(4)","営業利益(5)",
            "Food比率（目標25%）","Labor比率（目標35%）"]
    r = 4
    for key in keys:
        c = ws.cell(r, 1, key); c.font = Font(bold=True, color="FFFFFF"); c.fill = F_SEC
        for j in range(2, 16):
            ws.cell(r, j).fill = F_SEC
        r += 1
        h = ws.cell(r, 1, "店舗"); h.font = Font(bold=True, color="FFFFFF"); h.fill = F_HDR
        for j, m in enumerate(["年計"] + MONTHS, start=2):
            cc = ws.cell(r, j, m); cc.font = Font(bold=True, color="FFFFFF")
            cc.fill = F_HDR; cc.alignment = Alignment(horizontal="center")
        r += 1; first = r
        for tab in TABS:
            ws.cell(r, 1, tab).border = BORD
            src = RIDX[tab][key]
            cc = ws.cell(r, 2, f"='{tab}'!B{src}")
            cc.number_format = PCTFMT if "比率" in key else NUMFMT; cc.border = BORD
            for j, m in enumerate(MONTHS, start=3):
                c2 = ws.cell(r, j, f"='{tab}'!{MCOL[m]}{src}")
                c2.number_format = PCTFMT if "比率" in key else NUMFMT; c2.border = BORD
            r += 1
        if "比率" not in key:
            ws.cell(r, 1, "合計").font = Font(bold=True)
            for j in range(2, 15):
                cc = ws.cell(r, j, f"=SUM({get_column_letter(j)}{first}:{get_column_letter(j)}{r-1})")
                cc.font = Font(bold=True); cc.fill = F_YR
                cc.number_format = NUMFMT; cc.border = BORD
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 20
    for j in range(2, 16):
        ws.column_dimensions[get_column_letter(j)].width = 12


def new_wb(period="21期"):
    global PERIOD
    PERIOD = period
    wb = Workbook(); wb.remove(wb.active)
    build_summary(wb.create_sheet("全体サマリー"))
    for t in TABS:
        build_pl_tab(wb.create_sheet(t), t)
    return wb


if __name__ == "__main__":
    wb = new_wb(); wb.save("_skeleton.xlsx")
    for t in TABS:
        print(f"{t:<14} 行数 {len(LAYOUTS[t]):>3}  (固有行 {len(EXTRA_COGS[t])+len(EXTRA_SGA[t]):>2})")
