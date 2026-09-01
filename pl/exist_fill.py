#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""21期PLシートからの穴埋め

--- ★このシートの位置づけ（利用者確認 2026-08-27）------------------------
「22期からは既存PLは無視して大丈夫。そもそも、21期も無視していいよ。
  税理士が作ってるわけじゃなくて、俺が作ってたから。」

つまり既存の21期PLは【会計士の成果物ではなく、社長ご自身が手で作ったもの】。
正として扱うべき資料ではない。書類（請求書・銀行明細・カード明細）が
入っていない月を埋めておくための【暫定値】として使う。

    ・書類が入ったセルは、必ず書類の値が勝つ（このモジュールは空きセルしか埋めない）
    ・「既存PLとの食い違い」シートは “間違い探し” ではなく
      “書類で置き換えたところの記録” として読む
    ・22期では一切使わない（8月も使わない。NO_EXIST_MONTHS 参照）
    ・請求書フォルダを読み進めるほど、ここから来る値は減っていく

★ただし【検算の相手】としては引き続き使う。カード明細の合計や月ズレの検証で
  実際にこちらのバグを何度も見つけている（cards.py・transfers.py の検算）。
  「一致したら安心材料」「違っても書類が優先」という距離感で扱う。
"""
import exist_pl

MONTHS = exist_pl.MONTHS

# (タブ, 既存の行名) -> 新シートの行名
ALIAS = {
    ("さわら十三里屋", "仕入（Freeeカードリアル）"): "仕入（freeeカード）",
    ("さわら十三里屋", "消耗品費（Freeeカード）"): "消耗品費（freeeカード）",
    ("さわら十三里屋", "人件費"): "人件費（アルバイト）",   # 4〜7月がpayrollと完全一致
    ("さわら十三里屋", "広告宣伝費"): "広告宣伝費（共通宣伝費）",
    ("さわら十三里屋", "地代家賃"): "地代家賃（賃料）",
    ("りゅうちゃん", "売上（税込）"): "売上",
    ("りゅうちゃん", "仕入（freeeカード山中ストアー）"): "仕入（やまなか）",
    ("りゅうちゃん", "仕入（沖縄六角堂）"): "仕入（六角堂）",
    ("りゅうちゃん", "仕入（藤原ストアー）"): "仕入（藤原ストア）",
    ("りゅうちゃん", "仕入（平良洋酒店）"): "仕入（平洋酒店）",
    ("タコとハイボール", "人件費（社員）"): "人件費（店長）",
    ("業務課", "人件費（社員）"): "人件費（店長）",
    ("業務課", "地代家賃"): "地代家賃（賃料）",
    ("業務課", "外注費（SUN-X)"): "外注費（SUN-X）",        # 既存は括弧が半角
    ("業務課", "広告宣伝費"): "広告宣伝費（共通宣伝費）",
    ("焼きたて屋", "消費税#2"): "出前館消費税",   # 既存は「消費税」が2つある
    ("焼きたて屋", "広告宣伝費"): "広告宣伝費（共通宣伝費）",
    ("焼きたて屋", "耗品費"): "消耗品費",                   # 既存の誤記
    ("神栖横丁", "人件費（社員）"): "人件費（店長）",
    ("神栖横丁", "地代家賃（ともえ）"): "地代家賃（賃料）",
    ("神栖横丁", "電気"): "水道光熱費（電気料金）",
    ("神栖横丁", "水道"): "水道光熱費（水道料金）",
    ("神栖横丁", "ガス"): "水道光熱費（ガス料金）",
    ("神栖横丁", "ドリーム"): "居酒屋ドリーム",
    ("鳥害対策課", "人件費"): "人件費（店長）",
    ("鳥害対策課", "地代家賃"): "地代家賃（賃料）",
}

# 入れないもの。(タブ, 既存の行名) -> 理由
SKIP = {
    ("りゅうちゃん", "沖縄六角堂（8％対象）"):
        "内訳の作業用の行。8%1,097,585＋10%181,214＝1,278,799 で、"
        "「仕入（沖縄六角堂）」1,181,018 と合わない。どちらが正か決められない",
    ("りゅうちゃん", "沖縄六角堂（10％対象）"):
        "同上。「仕入（沖縄六角堂）」との関係がはっきりしない",
    ("りゅうちゃん", "〃"):
        "ディット記号だけの行。年計1,016,278。何を指すか決められない",
    ("りゅうちゃん", "〃#2"):
        "同上。「〃」の行がもう1つある",
    ("さわら十三里屋", "水道光熱費"):
        "中身は電気料金だけだった（10月49,577が新シートの電気と1円まで一致）。"
        "なめがたの請求書から11か月とも入っているので写す必要がない",
    ("業務課", "仕入（クリーン＆ケミカル）"): "全月0円",
    ("さわら十三里屋", "雑費"):
        "★シンクロエンターテイメントへの支払いと二重計上になる。既存PLは11月の"
        "「雑費」に230,730、新シートは銀行明細から12月の「その他経費」に230,709。"
        "銀行にシンクロへの支払いは2026/02/17の253,780（税込）1件しかなく、"
        "253,780÷1.1＝230,709。既存PLの雑費も年に11月の1件だけ。同じ支払いと見て"
        "既存PLのほうを写さない（新シートは書類の値を採る方針）。"
        "計上月は利用者判断で11月にそろえた（2026-08-21）。"
        "新シートは11月の「その他経費」に230,709で入っている",
}

# ★1セルだけ写さないもの。(タブ, 行, 月) -> 理由。SKIP は行まるごと、こちらは1か月だけ。
SKIP_CELL = {
    ("韓国酒場ハナ", "その他経費", "12月"):
        "既存21期PLに ＋54,200（プラス）で入っているが、横丁の12月分請求書"
        "に該当する項目が無く、カード明細・銀行明細にも見当たらない。"
        "何の費用か決められないので写さない。利用者判断「保留」（2026-08-23）",
    # ★2509月のタカギの請求書を読んだら、既存PLの2セルが少なかった（2026-09-01）。
    #   請求書は 税抜164,000（インスタ広告@80,000×2＝10%対象160,000 ＋ 交通費4,000＝対象外）。
    #   既存PLは 76,364×2＝152,728 で 11,272 少ない。
    #   利用者指示「交通費4,000は横丁で」→ 十三里屋80,000／横丁84,000。inv2509.py が入れる。
    ("さわら十三里屋", "タカギ（Instagramコンサル）", "9月"):
        "2509月の請求書（税抜164,000）で置き換えた。inv2509.py 参照",
    ("神栖横丁", "タカギ", "9月"):
        "2509月の請求書（税抜164,000）で置き換えた。交通費4,000はこちら。inv2509.py 参照",
    # ★2510月の藤原ストアーの請求書は空瓶・空樽代が −880（税込）だった。既存PLは −800。
    ("りゅうちゃん", "仕入（藤原ストアー値引き）", "10月"):
        "2510月の請求書（空瓶・空樽代金など −880税込 → 税抜−815）で置き換えた。inv2509.py 参照",
}
# ★神栖横丁「その他経費」の 9月〜12月 各2,560 は USEN NETWORKS の2契約だった
#   （購入企業ID 21298539／21298540。各 USEN NET 1,280＋消費税128＝税込1,408）。
#   利用者指示 2026-09-01「その他経費は通信費」→「通信費（有線ネット）」へ。
#   既存PLも6月・7月は同じものを「通信費（有線ネット）」に入れていた。
#     9月・10月・11月 … 請求書があるので請求書から入れる（inv2509.py / inv11.py）
#     12月           … まだ請求書を読んでいないので MOVE_CELL で行だけ移す
for _m in ("9月", "10月", "11月"):
    SKIP_CELL[("神栖横丁", "その他経費", _m)] = (
        "中身は USEN NETWORKS の2契約（各 USEN NET 1,280＋消費税128）。"
        "請求書から「通信費（有線ネット）」に入れ直した"
        "（利用者指示 2026-09-01「その他経費は通信費」）")

# 既存21期PLの値はそのまま使うが、入れる【行】だけ差し替えるもの。
#   (タブ, 新シートの行, 月) -> (移す先の行, 理由)
MOVE_CELL = {
    ("神栖横丁", "その他経費", "12月"): (
        "通信費（有線ネット）",
        "9月〜11月と同じ USEN NETWORKS の2契約 2,560 とみられる。"
        "2512月の請求書はまだ読んでいないので金額は既存21期PLのまま、行だけ移した"
        "（利用者指示 2026-09-01「その他経費は通信費」）"),
}

# 既存21期PLの金額を書き換えるもの。(タブ, PL行, 月) -> (直した額, 理由)
# ★2509月の請求書を読んで、既存PLが【税込のまま】入っていると分かったもの。
#   税抜に直す指示が出た月だけ、ここに1セルずつ書く（推測で広げない）。
VALUE_FIX = {}
_DTB = ("2509月の請求書で、既存PLが税込のまま入っていると分かった。"
        "請求書は COSMOSNET 3,000＋消費税300＝3,300。"
        "6月・7月は請求書から入れていて3,000なので、そちらが正しい形。"
        "利用者指示 2026-09-01「帝国データバンクが税抜きにして」")
for _m in ("9月", "10月", "11月", "12月", "1月", "3月", "4月", "5月"):
    VALUE_FIX[("本部", "帝国データバンク", _m)] = (3000, _DTB)
_SCPN = ("2509月の請求書（SPCNスプリーム）は "
         "『LINE公式アカウントサポート50,000／小計50,000／調整額4,000／"
         "合計（消費税10％込）54,000』。税込54,000として逆算すると49,091。"
         "利用者指示 2026-09-01「SCPNは税抜きにして」")
for _m in ("9月", "10月", "11月", "12月"):
    VALUE_FIX[("本部", "SCPN", _m)] = (49091, _SCPN)
# ★2月の帝国データバンクは 99,000 で、ほかの月（3,300）と桁が違う。
#   年会費など別物の可能性があるので触っていない。2602月の請求書を読むときに確かめる。
# --- 2510月の請求書で分かったぶん（2026-09-01）------------------------------
VALUE_FIX[("神栖横丁", "ウゴーク", "10月")] = (
    30000,
    "2510月の請求書（INV-0000000071・件名『2025年10月分』）は "
    "MEO対策業務手数料 1×30,000（税込33,000）だけだった。既存PLは80,000。"
    "9月は請求書も80,000で一致していたので、10月から下がったのに写し替えていない")
VALUE_FIX[("本部", "オフィスで野菜", "10月")] = (
    4167,
    "KOMPEITO『OFFICE DE YASAI 2025年10月分』は8%内税で税込4,500（内消費税333）。"
    "既存PLは税込のまま入っていた")
VALUE_FIX[("業務課", "仕入（江東微生物研究所）", "10月")] = (
    1800,
    "請求書（2510-22-03699）は検査料金1,800＋消費税180＝1,980。"
    "既存PLは税込のまま入っていた")
VALUE_FIX[("りゅうちゃん", "仕入（六角堂）", "9月")] = (
    152995,
    "2509月の請求書は税込165,769（10%対象29,425→26,750／8%対象136,344→126,245）＝152,995。"
    "既存PLは152,994で1円少なかった")
VALUE_FIX[("りゅうちゃん", "仕入（六角堂）", "10月")] = (
    122896,
    "2510月の請求書は税込133,019（10%対象16,049→14,590／8%対象116,970→108,306）＝122,896。"
    "既存PLは122,895で1円少なかった")
VALUE_FIX[("本部", "チムニータウン", "9月")] = (
    101619,
    "2509月の請求書（西野亮廣講演会 in 茨城県神栖市）は税抜101,619＋消費税10,161＝"
    "税込111,780。既存PLは101,618で1円少なかった。利用者指示 2026-09-01")

VALUE_FIX[("りゅうちゃん", "仕入（平洋酒店）", "11月")] = (
    42415,
    "2511月の請求書（平良洋酒店(有)の納品書・請求書）は "
    "『課税対象額40,115（10%）／消費税合計4,012／税込金額44,127』＋"
    "『課税額（県外出荷）2,300』＝合計46,427。既存PLはこの46,427（税込）を"
    "そのまま入れていた。県外出荷ぶんは酒税とみられ消費税の課税対象外なので、"
    "税抜は 40,115＋2,300＝42,415。利用者指示 2026-09-01「2.税抜きで」")

# 小計・利益・比率の行。新シートは数式で持っているので写さない。
# ★「期首|期末」を入れていたのは誤り（2026-08-21 に修正）。
#   期首棚卸し・期末棚卸しは計算行ではなく入力行で、新シートでも入力セル。
#   既存21期PLには5タブ×11か月ぶん入っていたのに、これで弾いてしまっていた。
#   売上原価(a) は数式なのでこちらは除外したままでよい。
import re
_CALC = re.compile(r"合計|利益|比率|\(\d+\)|＝|=|売上原価\(a\)")

# ★カード明細を費目別に割った月は、既存PLの一括額を写すと二重計上になる。
#   既存PLは1枚の請求額をまるごと「JCBカード」「三井住友カード」の1行に入れている。
#   新シートは cards.py が明細1行ずつを費目へ割り振る（利用者指示 2026-08-20）。
#   両方入れると同じ支払いを2回数えることになる。
#     実測: JCB 7月 … cards.py の税込合計 624,249 ＝ 既存PL「JCBカード」7月 624,249
#           1円まで一致した。同じ明細である裏づけ（2026-08-21）。
#   明細のある月だけを外す。明細の無い月（9月〜6月）は既存PLが唯一の情報源なので写す。
_LUMP = {"JCB": "JCBカード", "三井住友": "三井住友カード"}
# 一括で受けている行（カード会社ごとの請求額まるごと）。タブは本部だけ。
LUMP_ROWS = {"本部": list(_LUMP.values())}


def split_by_cards():
    """cards.py が費目別に割った (タブ, 一括行, 月) の集合。写してはいけないセル。"""
    import cards
    out = set()
    for tab, _merch, plrow, _ex, _tax, _src, m, _used, iss in cards.rows():
        if plrow is None:
            continue
        out.add((tab, _LUMP[iss], m))
    return out


def not_posted_rows():
    """わざと写さなかったセル (タブ, 行, 月, 既存PLの額, 理由) を列挙。

    明細ログに「載せなかった記録」として出し、セルのメモに理由が出るようにする。
    空欄を見て「なぜ入っていないのか」を探さずに済ませるため（2026-08-23）。
    """
    ex, _ = exist_pl.load()
    for tab, plrow, m in sorted(split_by_cards()):
        v = ex.get(tab, {}).get(plrow, {}).get(m)
        if not v:
            continue
        yield (tab, plrow, m, int(v),
               f"既存21期PLの「{plrow}」{m}は {int(v):,}円 の一括額。"
               "カード明細を費目別に割った月なので写していない。"
               "写すと同じ支払いを2回数えることになる（実測で税込1円まで一致）")
    for (tab, plrow, m), why in SKIP_CELL.items():
        v = ex.get(tab, {}).get(plrow, {}).get(m)
        yield (tab, plrow, m, int(v) if v else 0, why)


# ★8月（21期の最終月）は既存21期PLを使わない（利用者指示 2026-08-23）。
#   「8月は基本、CSVやPDFから全部数字入れてください。既存は無視してくださいね」
#   いまのところ会計士の既存PLに8月の値は1つも無いが、あとから入っても
#   写さないようにここで止める。足りないところは保留リストに出す。
NO_EXIST_MONTHS = {"8月"}


def _target(tab, item):
    """既存の行名 → 新シートの行名。写さないものは None。"""
    if (tab, item) in SKIP or _CALC.search(item):
        return None
    return ALIAS.get((tab, item), item)


# 1件この額以下（税抜）の接待交際費は会議費に回す（利用者指示 2026-08-22）。
# cards.py / engine.py と同じ扱いを、既存PLから写すぶんにも効かせる。
# ★既存PLの1セルは1件とは限らない（その月ぶんの合計）。
#   21期で当たったのは業務課9月の5,000だけで、これは1件ぶんと分かっている。
#   月に何件も入った合計が5,000円以下になる月が出てきたら、ここを見直すこと。
KAIGI_LIMIT = 5000


def rows(wb):
    """(タブ, PL行, 月, 値, 元, メモ) を列挙。空いているセルだけ。"""
    import build2
    ex, order = exist_pl.load()
    split = split_by_cards()
    for tab in order:
        if tab not in build2.RIDX or tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for item in order[tab]:
            plrow = _target(tab, item)
            if plrow is None or plrow not in build2.RIDX[tab]:
                continue
            for m in MONTHS:
                v = ex[tab][item].get(m)
                if not v:            # 未入力・0円は写さない
                    continue
                if m in NO_EXIST_MONTHS:
                    continue          # 書類だけで組む月（利用者指示 2026-08-23）
                if (tab, plrow, m) in split:   # カード明細を費目別に割った月
                    continue
                if (tab, item, m) in SKIP_CELL or (tab, plrow, m) in SKIP_CELL:
                    continue        # 1セルだけ写さないもの（SKIP_CELL）
                if (tab, plrow, m) in VALUE_FIX:
                    v = VALUE_FIX[(tab, plrow, m)][0]   # 税込→税抜などの直し
                moved = MOVE_CELL.get((tab, plrow, m))
                if moved:                                # 行だけ差し替える
                    if moved[0] not in build2.RIDX[tab]:
                        continue
                    plrow = moved[0]
                c = ws[f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
                # ★書類から入っている。絶対に上書きしない。
                #   0 も「書類を足したらちょうど0になった」という結論なので守る。
                #   例: 韓国酒場ハナ 6月のその他経費は
                #       ChatGPT +2,727（freee）と横丁のその他値引き ▲2,727 で 0。
                #       `if c.value:` だと0を未入力と見て既存PLの▲2,727を写してしまい、
                #       ChatGPTの費用が消えていた（2026-08-23 に直した）。
                if c.value is not None:
                    continue
                note = "既存21期PLから転記"
                if moved:
                    note += f"（★もとの行「{item}」から移した: {moved[1]}）"
                elif (tab, plrow, m) in VALUE_FIX:
                    note += f"（★{VALUE_FIX[(tab, plrow, m)][1]}）"
                if plrow != item:
                    note += f"（既存の行名「{item}」）"
                if plrow == "接待交際費" and int(v) <= KAIGI_LIMIT:
                    plrow = "会議費"
                    note += "／5,000円以下なので会議費へ（利用者指示 2026-08-22）"
                    if plrow not in build2.RIDX[tab]:
                        continue
                yield (tab, plrow, m, int(v), f"既存21期PL {tab}", note)


def conflicts(wb):
    """既存PLと新シートで金額が食い違うセル。(タブ, 行, 月, 新, 既存)"""
    import build2
    ex, order = exist_pl.load()
    split = split_by_cards()
    out = []
    for tab in order:
        if tab not in build2.RIDX or tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for item in order[tab]:
            plrow = _target(tab, item)
            if plrow is None or plrow not in build2.RIDX[tab]:
                continue
            for m in MONTHS:
                if m in NO_EXIST_MONTHS:
                    continue          # 書類だけで組む月（利用者指示 2026-08-23）
                if (tab, plrow, m) in split:
                    continue
                if (tab, item, m) in SKIP_CELL or (tab, plrow, m) in SKIP_CELL:
                    continue        # 1セルだけ写さないもの（SKIP_CELL）
                v = ex[tab][item].get(m)
                if (tab, plrow, m) in VALUE_FIX:
                    v = VALUE_FIX[(tab, plrow, m)][0]   # 税込→税抜などの直し
                pl = plrow
                if (tab, plrow, m) in MOVE_CELL:        # 行だけ差し替えたもの
                    pl = MOVE_CELL[(tab, plrow, m)][0]
                    if pl not in build2.RIDX[tab]:
                        continue
                if pl == "接待交際費" and v and int(v) <= KAIGI_LIMIT:
                    pl = "会議費"                       # rows() と同じ読み替え
                    if pl not in build2.RIDX[tab]:
                        continue
                c = ws[f"{build2.MCOL[m]}{build2.RIDX[tab][pl]}"].value
                if not v or not c or isinstance(c, str):
                    continue
                if int(c) != int(v):
                    out.append((tab, plrow, m, int(c), int(v)))
    return out


def lump_months(wb):
    """一括行のまま残っている (タブ, 行, 月, 金額)。明細が入れば費目別にできる。"""
    import build2
    split = split_by_cards()
    out = []
    for tab, rows in LUMP_ROWS.items():
        if tab not in wb.sheetnames:
            continue
        for plrow in rows:
            r = build2.RIDX[tab].get(plrow)
            if r is None:
                continue
            for m in MONTHS:
                if m in NO_EXIST_MONTHS:
                    continue          # 書類だけで組む月（利用者指示 2026-08-23）
                if (tab, plrow, m) in split:
                    continue
                if (tab, plrow, m) in SKIP_CELL:
                    continue        # 1セルだけ写さないもの（SKIP_CELL）
                v = wb[tab][f"{build2.MCOL[m]}{r}"].value
                if v and not isinstance(v, str):
                    out.append((tab, plrow, m, int(v)))
    return out


def check(wb=None):
    """対応づけの取りこぼしが無いかを見る。"""
    import build2
    ex, order = exist_pl.load()
    lost = []
    for tab in order:
        if tab not in build2.RIDX:
            continue
        for item in order[tab]:
            if not any(ex[tab][item].values()):
                continue
            plrow = _target(tab, item)
            if plrow is None:
                continue
            if plrow not in build2.RIDX[tab]:
                lost.append((tab, item, sum(ex[tab][item].values())))
    assert not lost, ("既存PLに値があるのに新シートに行が無い（ALIAS か SKIP か "
                      "build2.py の行追加が要る）:\n  " +
                      "\n  ".join(f"{t} 「{i}」 {s:,}円" for t, i, s in lost))
    if wb is None:
        return
    # ★カード明細のルール（利用者指示 2026-08-21、2026-08-22 に見直し）:
    #   明細のある月は、取引先マスタで割れたぶんを費目別に、割れなかったぶんを
    #   カード会社の一括行に入れる。既存PLの一括額は写さない（rows() で除外済み）。
    #   （最初は「一括行が空であること」を見ていたが、未分類ぶんを一括行に
    #     残す作りにしたので前提が変わった）
    #   ここでは一括行が「未分類の合計」と一致することを見る。
    #   これが合っていれば、既存PLの額が紛れ込んでいないことも同時に確かめられる。
    import cards
    import collections
    unclassified = collections.Counter()
    for _t, _mc, plrow_, ex_, _tx, _sr, m_, _u, iss_ in cards.rows():
        if plrow_ is None:
            unclassified[(_LUMP[iss_], m_)] += ex_
    bad = []
    for tab, plrow, m in sorted(split_by_cards()):
        r = build2.RIDX[tab].get(plrow)
        if r is None:
            continue
        v = wb[tab][f"{build2.MCOL[m]}{r}"].value
        got = int(v) if v and not isinstance(v, str) else 0
        want = unclassified.get((plrow, m), 0)
        if got != want:
            bad.append(f"{tab} 「{plrow}」{m}: シート {got:,} ／ 未分類の合計 {want:,}")
    assert not bad, ("カード一括行が未分類の合計と合わない"
                     "（既存PLの額が紛れ込んでいるか、取りこぼし）:\n  "
                     + "\n  ".join(bad))

    # ★棚卸しの連鎖: 前月の期末＝当月の期首。崩れると売上原価が狂う。
    #   売上原価(a) = 仕入合計(2) + 期首棚卸し - 期末棚卸し なので、
    #   ここがずれると売上総利益から下が全部おかしくなる。
    chain = []
    for tab in build2.RIDX:
        if tab not in wb.sheetnames:
            continue
        rs = build2.RIDX[tab].get("期首棚卸し"), build2.RIDX[tab].get("期末棚卸し")
        if None in rs:
            continue
        for i in range(len(MONTHS) - 1):
            m, nx = MONTHS[i], MONTHS[i + 1]
            end = wb[tab][f"{build2.MCOL[m]}{rs[1]}"].value
            beg = wb[tab][f"{build2.MCOL[nx]}{rs[0]}"].value
            if not end or not beg or isinstance(end, str) or isinstance(beg, str):
                continue
            if int(end) != int(beg):
                chain.append(f"{tab}: {m}の期末{int(end):,} ≠ {nx}の期首{int(beg):,}")
    assert not chain, "棚卸しの連鎖が切れている:\n  " + "\n  ".join(chain)


def hold_rows(wb=None):
    """保留リストへ出すもの。"""
    ex, _ = exist_pl.load()
    out = []
    if wb is not None:
        for tab, plrow, m, v in lump_months(wb):
            out.append((m, tab, f"{plrow}（一括のまま {v:,}円）",
                        "カード明細がまだ無い月。既存21期PLの請求額まるごとを暫定で置いている。"
                        "明細（JCBは20xxxxmeisai.csv、三井住友は20xxxx.csv）を "
                        "※請求書※/freeeカード明細/21期/ に入れてもらえれば、"
                        "cards.py が費目別に割り直してこの行は空になる"))
    for (tab, plrow, m), why in SKIP_CELL.items():
        v = ex.get(tab, {}).get(plrow, {}).get(m)
        out.append((m, tab, f"既存PL「{plrow}」{m}"
                    + (f"（{int(v):,}円）" if v else ""), why))
    return out + _skip_holds(ex)


def _skip_holds(ex):
    out = []
    for tab, plrow, m in sorted(split_by_cards()):
        v = ex.get(tab, {}).get(plrow, {}).get(m)
        if not v:
            continue
        out.append((m, tab, f"既存PL「{plrow}」{m}（{v:,}円）",
                    "カード明細を費目別に割った月なので、既存PLの一括額は写していない。"
                    "写すと同じ支払いを2回数えることになる（実測で税込1円まで一致）"))
    for (tab, item), why in SKIP.items():
        s = sum(ex.get(tab, {}).get(item, {}).values())
        if not s:
            continue
        out.append(("", tab, f"既存PL「{item}」（年計{s:,}円）", why))
    return out


if __name__ == "__main__":
    import warnings; warnings.filterwarnings("ignore")
    import openpyxl, collections
    check()
    wb = openpyxl.load_workbook("損益計算書_21期テスト版.xlsx")
    rs = list(rows(wb))
    by = collections.Counter()
    amt = collections.Counter()
    for tab, _r, _m, v, _s, _n in rs:
        by[tab] += 1; amt[tab] += v
    print(f"{'タブ':<14}{'埋めるセル':>10}{'金額':>16}")
    print("-" * 42)
    for tab in by:
        print(f"{tab:<14}{by[tab]:>10}{amt[tab]:>16,}")
    print("-" * 42)
    print(f"{'合計':<14}{sum(by.values()):>10}{sum(amt.values()):>16,}")
    print(f"\n食い違い（新シート優先。上書きしない）: {len(conflicts(wb))}セル")
    print(f"保留へ出すもの: {len(hold_rows())}件")
