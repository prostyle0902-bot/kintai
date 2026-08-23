#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""確定データを v2 ワークブックへ転記"""
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side
import build2, sales, inv6, inv7, payroll, cards, board, demaekan, kameya, yokocho, fixed_costs, shokaihi
import rikuji, eneos, yokocho_bank, store_bank, transfers
import namefa, shiina, exist_fill, inv8, nihonshokken, norow, cellnote, status8

STAMP = "2026-08-20 06:40"

# 店舗ごとに行名が違うものの読み替え（既存スプシに合わせる）
REMAP = {("韓国酒場ハナ", "仕入（やまなか）"): "仕入（山中ストアー）"}
F_POST = PatternFill("solid", fgColor="D9F2D0")
THIN = Side(style="thin", color="B4C6E7")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---- 請求書 2607月（＝7月計上）: (タブ, 取引先, PL行, 税抜, 消費税, 元ファイル, 備考) ----
INV = [
 ("りゅうちゃん","片野商店","水道光熱費（ガス料金）",19583,1958,"りゅうちゃん/片野商店　りゅうちゃん.pdf",
  "PDF2ページ＝請求書2枚を合算（⑩17,691＋⑨1,892）／既存スプシ19,583と一致"),
 ("韓国酒場ハナ","片野商店","水道光熱費（ガス料金）",5518,551,"ハナ/片野商店　ハナ.pdf",""),
 ("もも焼きJAPAN","片野商店","水道光熱費（ガス料金）",7738,773,"もも焼き/片野商店　もも焼き.pdf",""),
 ("タコとハイボール","片野商店","水道光熱費（ガス料金）",8145,814,"タコハイ/片野商店　タコハイ.pdf",
  "既存スプシ8,145と一致"),
 ("神栖横丁","片野商店","水道光熱費（ガス料金）",19245,1924,"横丁/片野商店　横丁.pdf",""),
 ("神栖横丁","水道料金","水道光熱費（水道料金）",76810,7681,"横丁/水道料金　横丁振替.pdf",
  "水道50,655＋下水道33,836（7月分）"),
 ("神栖横丁","NTTファイナンス","通信費",7534,753,"横丁/NTTファイアンス　横丁振替.pdf",
  "ファイル名「ファイアンス」は誤記。マスタのNTTファイナンスと同一（確認済）"),
 ("神栖横丁","USEN","通信費（USEN）",1280,128,"横丁/USEN 横丁振替.pdf","請求書No F-1-20260804-104412-01"),
 ("神栖横丁","USEN","通信費（USEN）",1280,128,"横丁/USEN　横丁振替.pdf","請求書No F-1-20260804-105743-01（別契約）"),
 ("神栖横丁","アルソック","ALSOK",20500,2050,"横丁/アルソック　横丁振替.pdf","既存スプシ20,500と一致"),
 ("神栖横丁","ウゴーク","ウゴーク",30000,3000,"横丁/ウゴーク　横丁8月末.pdf","既存スプシ30,000と一致"),
 ("神栖横丁","業務（グリスト清掃）","グリスト清掃業務課",40000,4000,"横丁/業務　横丁.pdf",
  "既存スプシ グリスト清掃業務課40,000と一致"),
 ("焼きたて屋","門倉石油","水道光熱費（ガス料金）",21122,2112,"焼きたて屋/門倉石油　焼きたて屋.pdf",
  "既存スプシ21,122と一致"),
 ("焼きたて屋","丸善エコアース","ごみ処分費",3600,360,"焼きたて屋/丸善エコアース　焼きたて屋.pdf",
  "既存スプシ ごみ処分費3,600と一致"),
 ("さわら十三里屋","椎名環境整備","廃棄物処分",12000,1200,"十三里屋/椎名環境整備　十三里屋8月末.pdf",
  "既存スプシ12,000と一致"),
 # なめがたしろはとふぁーむ（1枚を既存スプシと同じ行に分解）
 ("さわら十三里屋","なめがたしろはとふぁーむ","仕入（なめファ8％）",316948,25355,
  "十三里屋/なめがたしろはとふぁーむ　十三里屋8月末.pdf","仕入食材7月（8%軽減）／既存スプシ316,948と一致"),
 ("さわら十三里屋","なめがたしろはとふぁーむ","仕入（なめファ10％）",4600,460,
  "十三里屋/なめがたしろはとふぁーむ　十三里屋8月末.pdf","仕入れ資材7月／既存スプシ4,600と一致"),
 ("さわら十三里屋","なめがたしろはとふぁーむ","地代家賃（賃料）",200000,20000,
  "十三里屋/なめがたしろはとふぁーむ　十三里屋8月末.pdf","家賃"),
 ("さわら十三里屋","なめがたしろはとふぁーむ","水道光熱費（水道料金）",1100,110,
  "十三里屋/なめがたしろはとふぁーむ　十三里屋8月末.pdf","水道代7月分"),
 ("さわら十三里屋","なめがたしろはとふぁーむ","水道光熱費（電気料金）",54436,5443,
  "十三里屋/なめがたしろはとふぁーむ　十三里屋8月末.pdf","佐原従量電灯C 22,851＋佐原低圧電力 31,585"),
 ("さわら十三里屋","なめがたしろはとふぁーむ","その他経費（ロイヤリティ）",98943,9894,
  "十三里屋/なめがたしろはとふぁーむ　十三里屋8月末.pdf",
  "ブランド使用3.0% 31,245＋本部管理経費6.5% 67,698／既存スプシ98,943と一致"),
 # ヴィーナスダイニング（タコハイで確定）
 ("タコとハイボール","ヴィーナスダイニング","仕入（インフォマート8％）",126455,10116,
  "タコハイ/ヴィーナスダイニング　タコハイ8月末.pdf","材料代8%軽減／既存スプシ126,455と一致"),
 ("タコとハイボール","ヴィーナスダイニング","仕入（インフォマート10％）",21400,2140,
  "タコハイ/ヴィーナスダイニング　タコハイ8月末.pdf","材料代10%／既存スプシ21,400と一致"),
 ("タコとハイボール","ヴィーナスダイニング","仕入（インフォマート利用料）",2500,250,
  "タコハイ/ヴィーナスダイニング　タコハイ8月末.pdf","既存スプシ2,500と一致"),
 ("タコとハイボール","ヴィーナスダイニング","ロイヤリティ",45311,4531,
  "タコハイ/ヴィーナスダイニング　タコハイ8月末.pdf","ロイヤリティ5%／既存スプシ45,311と一致"),
 # 日本食研（1枚を3店舗に分割）
 ("韓国酒場ハナ","日本食研","仕入（日本食研）",31515,2521,
  "ハナ/日本食研　ハナ・もも焼き・タコハイ8月末.pdf","内訳(8795207)韓国酒場ハナ／全品8%軽減"),
 ("もも焼きJAPAN","日本食研","仕入（日本食研）",6087,487,
  "ハナ/日本食研　ハナ・もも焼き・タコハイ8月末.pdf","内訳(8489343)もも焼JAPAN／全品8%軽減"),
 ("タコとハイボール","日本食研","仕入（日本食研）",5056,404,
  "ハナ/日本食研　ハナ・もも焼き・タコハイ8月末.pdf",
  "内訳(9278338)神栖横丁5丁目＝タコハイ店／既存スプシ5,056と一致"),
 # 鳥害対策課
 ("鳥害対策課","日本鳩対策センター","仕入（日本鳩対策センター）",1003398,100340,
  "鳥害対策/日本鳩対策センター.pdf","税抜御買上額／既存スプシ1,003,398と一致"),
 ("鳥害対策課","西尾レントオール","外注費（西尾レントオール）",247350,24735,
  "鳥害対策/西尾レントオール.pdf","10%対象計"),
 # 業務課
 ("業務課","ビーエム","仕入（ビーエム）",129580,12958,"業務/ビーエム　.pdf",
  "当月御買上額（繰越330,922は含めない）／既存スプシ129,580と一致"),
 ("業務課","ニッセーデリカ","仕入（ニッセーデリカ）",2080,208,"業務/ニッセーデリカ.pdf",
  "既存スプシは税込2,288。新シートは税抜2,080で統一"),
 ("業務課","江東微生物研究所","仕入（江東微生物研究所）",600,60,"業務/江東微生物研究所.pdf",
  "既存スプシは税込660。新シートは税抜600で統一"),
 ("業務課","ケミカルテクノロジー","仕入（ケミカルテクノロジー）",72900,7290,"業務/ケミカルテクノロジー.pdf",
  "既存スプシ72,900と一致"),
 ("業務課","エス・アイ・ビー・エス","外注費（SIBS）",66000,6600,"業務/エスアイビーエス.pdf",
  "既存スプシ66,000と一致"),
 ("業務課","SUN-X","外注費（SUN-X）",36000,3600,"業務/SUN-X.pdf","既存スプシ36,000と一致"),
 ("業務課","ピカピカ","外注費（ピカピカ）",45455,4545,"業務/ピカピカ　7月末振替.pdf",
  "税込50,000・うち消費税4,545／既存スプシ45,455と一致"),
 ("業務課","ミノアカ","外注費（ミノアカ）",275000,27500,"業務/ミノアカ.xlsx",
  "7月分12件（信金各支店・ハウスクリーニング等）／既存スプシ275,000と一致"),
 # ★日本ビルメン（JBQ）の会費は本部の「JBQ」行に fixed_costs.py が毎月20,000を
 #   入れている。銀行の支払いも月1回だけ（千葉銀行9-10月20,385／PayPay11月〜20,000）。
 #   ここに入れると二重計上になるので外した（2026-08-21）。
 ("業務課","リペアボックス","修繕費（リペアボックス）",15000,1500,"業務/リペアボックス.pdf",
  "東京食堂（神栖市）修理／新規項目"),
 ("本部","帝国データバンク","帝国データバンク",3000,300,"業務/帝国データバンク.pdf",
  "COSMOSNET 2026.07利用分／新規項目"),
]

INV += inv7.INV7_ADD   # 2026-08-20 追加投入分（本部フォルダ8件を含む28ファイル）

INV_HOLD = []   # 2026-08-18 利用者指示により全件解決（行名＝取引先名）

# 取引先マスタに無い等で入れられないもの（タブ, 月, 取引先, 税込, 理由）
EXTRA_HOLD = []

# 金額のない実績データ（明細ログにのみ記録）
JISSEKI = [("焼きたて屋","丸善エコアース","焼きたて屋/丸善エコアース実績.pdf",
            "2026年7月 回収実績: 月間回収量80kg／回収袋数7袋（可燃・生ごみ）。"
            "金額は『ごみ処分費』3,600円として別途計上済み")]


def main(dst="損益計算書_21期テスト版.xlsx"):
    wb = build2.new_wb()
    ok = pd.read_csv("out_meisai.csv", encoding="utf-8-sig")
    try:
        hold = pd.read_csv("out_hold.csv", encoding="utf-8-sig")
    except pd.errors.EmptyDataError:          # 保留ゼロ件のとき
        hold = pd.DataFrame(columns=["利用日", "店舗", "取引先", "税込", "理由"])
    live = ok[ok["判定"] != "除外"].copy()
    live["税抜"] = live["税抜"].astype(int)

    posted, missing = 0, []
    for (tab, plrow, month), val in live.groupby(["店舗", "PL行", "計上月"])["税抜"].sum().items():
        plrow = REMAP.get((tab, plrow), plrow)
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "カード")); continue
        c = wb[tab][f"{build2.MCOL[month]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + int(val); c.fill = F_POST; c.number_format = build2.NUMFMT
        posted += 1
    for tab, vendor, plrow, ex, tax, src, biko in INV:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "請求書")); continue
        c = wb[tab][f"{build2.MCOL['7月']}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + ex; c.fill = F_POST; c.number_format = build2.NUMFMT
        posted += 1
    # 2608月 → 8月列（21期の最終月。届いたぶんから順に入れていく）
    inv8.check(wb)
    for tab, vendor, plrow, ex, tax, src, biko in inv8.INV8:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "請求書8月")); continue
        c = wb[tab][f"{build2.MCOL['8月']}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + ex; c.fill = F_POST; c.number_format = build2.NUMFMT
        posted += 1

    # 2606月 → 6月列（仕様どおり）
    for tab, vendor, plrow, ex, tax, src, biko in inv6.INV6:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "請求書6月")); continue
        c = wb[tab][f"{build2.MCOL['6月']}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + ex; c.fill = F_POST; c.number_format = build2.NUMFMT
        posted += 1

    # ===== JCB / 三井住友カード（支払日ベース） =====
    F_CARD = PatternFill("solid", fgColor="FFF2CC")
    card_cells, card_hold = 0, []
    # ★分類できなかったぶんは捨てず、カード会社の一括行に残す（2026-08-22）。
    #   明細を全部読むようにしたら187件・約302万円が取引先マスタに無かった。
    #   落とすと明細の総額とPLが合わなくなる。一括行に寄せておけば
    #   「費目別に割れたぶん＋未分類＝明細の総額」が常に保たれ、
    #   マスタに足すたびに一括行が減っていく。全件を保留リストにも出す。
    LUMP = {"JCB": "JCBカード", "三井住友": "三井住友カード"}
    agg = {}
    for tab, merch, plrow, ex, tax, src, m, used, iss in cards.rows():
        if plrow is None:
            card_hold.append((iss, merch, ex, used))
            plrow = LUMP[iss]
        agg[(tab, plrow, m)] = agg.get((tab, plrow, m), 0) + ex
    for (tab, plrow, m), val in agg.items():
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "カード(JCB/三井住友)")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_CARD; c.number_format = build2.NUMFMT
        card_cells += 1
    print("JCB/三井住友セル", card_cells, "／計", f"{sum(agg.values()):,}",
          "／未分類（一括行に残した）", len(card_hold), "件",
          f"{sum(x[2] for x in card_hold):,}円")

    # ===== board（売掛）→ 業務課・鳥害対策課・神栖横丁の売上 =====
    F_BOARD = PatternFill("solid", fgColor="FCE4EC")
    board_rows = list(board.rows())
    for tab, plrow, m, ex, tax, n, src, detail in board_rows:
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + ex; c.fill = F_BOARD; c.number_format = build2.NUMFMT
    print("boardセル", len(board_rows))
    # boardが正になる (タブ, 行, 月) — 既存スプシの売上転記から除外する
    board_owned = set(board.SUPPRESS) | {(t, r, m) for t, r, m, *_ in board_rows}

    # ===== 出前館の手数料（21期は手数料のみ／利用者判断 B-2） =====
    demaekan.check(sales)          # ①が既存PLと一致することを確認
    F_DEMAE = PatternFill("solid", fgColor="E2EFDA")
    demae_rows = list(demaekan.fee_rows())
    for tab, plrow, m, val, src, _ in demae_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "出前館")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_DEMAE; c.number_format = build2.NUMFMT
    demae_refund = list(demaekan.refund_rows())
    for tab, plrow, m, val, src, _ in demae_refund:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "出前館返金")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_DEMAE; c.number_format = build2.NUMFMT
    print("出前館セル", len(demae_rows) + len(demae_refund),
          "／手数料", f"{sum(x[3] for x in demae_rows):,}",
          "／返金", f"{sum(x[3] for x in demae_refund):,}")

    # ===== かめや（焼きたて屋のFC本部）=====
    kameya.check()                 # ロイヤリティの式が既存PLと合うことを確認
    F_KAME = PatternFill("solid", fgColor="FFF2CC")
    kame_rows = list(kameya.rows())
    for tab, plrow, m, val, src in kame_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "かめや")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_KAME; c.number_format = build2.NUMFMT
    kame_cash = list(kameya.cash_rows())
    for tab, plrow, m, val, src, note in kame_cash:
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_KAME; c.number_format = build2.NUMFMT
    print("かめやセル", len(kame_rows), "／現金過不足→雑損失", len(kame_cash),
          "セル 計", f"{sum(x[3] for x in kame_cash):+,}")

    # 横丁の請求書だけが埋める行。ほかの元データと重なったら事故なので止める。
    YOKO_EXCLUSIVE = {"地代家賃（賃料）", "地代家賃（駐車場利用料）", "地代家賃（共益費）",
                      "広告宣伝費（共通宣伝費）", "水道光熱費（電気料金）", "水道光熱費（水道料金）"}
    # ===== 神栖横丁 → 入居店舗への社内請求（21期10か月ぶん）=====
    # 請求書PDF40枚（10か月×4店舗）を yokocho_parse.py が読み、
    # yokocho_data.py に起こしてある。2026-08-22 に7月だけ→10か月に増やした。
    # ★このブロックは fixed_costs より前に置くこと。
    #   fixed_costs は「埋まっているセルは飛ばす」ので、請求書のほうが優先される。
    #   飛ばしたぶんは fixed_costs.check(wb) が定額と一致するか検算してくれる。
    yokocho.check(wb)
    F_YOKO = PatternFill("solid", fgColor="DDEBF7")
    yoko_rows = list(yokocho.rows())
    for tab, plrow, m, val, src in yoko_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "横丁社内請求")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        # ★家賃・駐車場・共益費・共通販促費・電気・水道は横丁の請求書だけが持つ行。
        #   ここに来る前に埋まっていたら二重計上なので止める
        #   （inv6.py が6月ぶんを持っていたのを 2026-08-22 に外した）。
        #   その他経費・消耗品費は他の元データも入る行なので足しこむ
        #   （例: 韓国酒場ハナ 6月の ChatGPT 2,727 はカード明細から）。
        if plrow in YOKO_EXCLUSIVE:
            assert not c.value, (f"{tab} {plrow} {m} に既に {c.value} が入っている。"
                                 f"横丁の請求書 {val:,} を足すと二重計上になる")
        c.value = int(c.value or 0) + val
        c.fill = F_YOKO; c.number_format = build2.NUMFMT
    print("横丁社内請求セル", len(yoko_rows), "／計",
          f"{sum(x[3] for x in yoko_rows):,}",
          "／", len({x[2] for x in yoko_rows}), "か月ぶん")

    # ===== 神栖横丁の口座振替（電気・ガス・水道・電話・USENほか）=====
    # 請求書が2606月・2607月しか無い費目を、横丁の口座の引落から埋める。
    # 行の割り当ては利用者指示（2026-08-21）。請求書から入っている取引先は飛ばす。
    yb_bad = yokocho_bank.check()
    F_YBK = PatternFill("solid", fgColor="E7E6E6")
    yb_rows = list(yokocho_bank.rows(wb))
    for tab, plrow, m, val, src, note in yb_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "横丁口座")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_YBK; c.number_format = build2.NUMFMT
    print("横丁口座振替セル", len(yb_rows), "／計", f"{sum(x[3] for x in yb_rows):,}",
          "／請求書と食い違った月", len(yb_bad))

    # ===== 店舗口座の口座振替（ガス代・鈴喜）=====
    # 店舗ごとに口座が分かれている。ただし11月から支払いがPayPay銀行に移ったので、
    # 店舗口座に毎月残っているのは口座振替だけ。それだけを入れる。
    sb_bad = store_bank.check(wb)
    sb_rows = list(store_bank.rows(wb))
    for tab, plrow, m, val, src, note in sb_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "店舗口座")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_YBK; c.number_format = build2.NUMFMT
    print("店舗口座振替セル", len(sb_rows), "／計", f"{sum(x[3] for x in sb_rows):,}",
          "／請求書と食い違った月", len(sb_bad))

    # ===== 仕入先への振込（千葉銀行の店舗口座＋PayPay銀行）=====
    # 2025年11月から支払いがPayPay銀行に移っているので、両方つないで1年ぶんにする。
    # 月ズレは支払日で決まる（月末払い＝1か月前／10日払い＝2か月前）。実測済み。
    transfers.check(wb)
    tr_rows = list(transfers.rows(wb))
    for tab, plrow, m, val, src, note in tr_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "振込")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_YBK; c.number_format = build2.NUMFMT
    print("振込セル", len(tr_rows), "／計", f"{sum(x[3] for x in tr_rows):,}")

    # ===== なめがたしろはとファーム → さわら十三里屋（請求書11枚）=====
    # 銀行だけだと年816万円の一括入金で8%/10%に割れず、いちばん大きな保留だった。
    # 請求書には家賃・ロイヤリティ・水道・電気まで入っているので6行が埋まる。
    # 6月分は inv6.py、7月分は下の INV が持っているので、ここは9月〜5月だけ。
    namefa.check(wb)               # 請求書の内部整合・銀行の振込額・書き込み先が空か
    nf_rows = list(namefa.rows())
    for tab, plrow, m, val, src, note in nf_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "なめがた")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_POST; c.number_format = build2.NUMFMT
    print("なめがたセル", len(nf_rows), "／計", f"{sum(x[3] for x in nf_rows):,}")

    # ===== 日本食研（1枚の請求書を店舗別に割る）=====
    # 1回の振込に複数店ぶんがまとまっていて銀行では割れなかったが、
    # 請求書に店舗ごとの【店計】があった（利用者指示 2026-08-22）。
    # 6月・7月は inv6.py / 下の INV が先に入れているので9月〜5月だけ。
    nihonshokken.check()
    ns_rows = list(nihonshokken.rows())
    for tab, plrow, m, val, src, note in ns_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "日本食研")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_POST; c.number_format = build2.NUMFMT
    print("日本食研セル", len(ns_rows), "／計", f"{sum(x[3] for x in ns_rows):,}")

    # ===== 受け皿の行が無かった3件（新設した行へ）=====
    # 利用者指示 2026-08-22「行を作成して」。
    #   もも焼きJAPAN「外注費（MEG design office）」… 請求書3枚（イベントサポート業務）
    #   もも焼きJAPAN「ヒラノ　タカシ」            … 請求書が無く費目を決められない
    #   タコとハイボール「共済掛金（SMBC）」        … 口座振替1件・非課税
    # ★千葉銀行のWEB振込額には振込手数料が乗っていることが今回わかった。
    #   norow.py は手数料を除いた本体だけを入れている。
    norow.check(wb)                # 銀行明細の実額・請求書の内部整合・書き込み先が空か
    nr_rows = list(norow.rows())
    for tab, plrow, m, val, src, note in nr_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "行を新設")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_POST; c.number_format = build2.NUMFMT
    print("新設行セル", len(nr_rows), "／計", f"{sum(x[3] for x in nr_rows):,}")

    # ===== 椎名環境整備の産廃処分（大口2件）→ 本部 =====
    # 銀行に242,000（2025/11）と221,100（2026/05）の大きな支払いがあり、
    # 毎月の13,200とは桁が違った。業務フォルダに請求書があり、場所は
    # 「本社倉庫」「事務所」。利用者判断で本部に入れる（2026-08-21）。
    shiina.check(wb)               # 請求書の内部整合・振込額・書き込み先が空か
    sh_rows = list(shiina.rows())
    for tab, plrow, m, val, src, note in sh_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "産廃")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_POST; c.number_format = build2.NUMFMT
    print("産廃セル", len(sh_rows), "／計", f"{sum(x[3] for x in sh_rows):,}")

    # ===== 業務課の車両費（請求書あり・11か月まとめて）=====
    # 陸事総合＝ETC高速代（請求書PDF）／ENEOS＝給油代（請求書CSV）。
    # どちらも既存PLが税込・税抜バラバラで、月ズレもあった。ここで全月入れ替える。
    rikuji.check()                 # PDF実物・銀行の引落・CP請求鑑CSVと突き合わせ
    eneos.check()
    F_CAR = PatternFill("solid", fgColor="FCE4D6")
    car_rows = list(rikuji.rows()) + list(eneos.rows())
    for tab, plrow, m, val, src, note in car_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "車両費")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_CAR; c.number_format = build2.NUMFMT
    print("陸事総合＋ENEOSセル", len(car_rows),
          "／陸事総合", f"{sum(x[3] for x in rikuji.rows()):,}",
          "／ENEOS", f"{sum(x[3] for x in eneos.rows()):,}")

    # ===== 毎月定額の自動引落・自動振込（請求書なし・全タブ）=====
    # ★請求書・カード明細・横丁の社内請求より後に置くこと。check() が
    #   「空であること」を見て二重計上を止めるので、先に置くと素通りする。
    fix_same = fixed_costs.check(wb)     # 重なったセルの金額一致を確認
    F_FIX = PatternFill("solid", fgColor="FFF2F2")
    fix_rows = list(fixed_costs.rows(wb))
    for tab, plrow, m, val, src, note in fix_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "本部定額")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_FIX; c.number_format = build2.NUMFMT
    print("定額セル", len(fix_rows), "／計", f"{sum(x[3] for x in fix_rows):,}",
          "／請求書と重なって飛ばしたセル", len(fix_same), "（金額は全部一致）")

    # ===== 本部の諸会費（会費類の受け皿。2026-08-20 新設）=====
    shokaihi.check()
    kaihi_rows = list(shokaihi.rows())
    for tab, plrow, m, val, src, note in kaihi_rows:
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "諸会費")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_POST; c.number_format = build2.NUMFMT
    print("諸会費セル", len(kaihi_rows), "／計", f"{sum(x[3] for x in kaihi_rows):,}")
    # かめやが正になる (タブ, 行, 月)。既存スプシからの売上転記より優先する。
    # ★これが無いと、あとの売上ループが同じセルを上書きして順番依存になる。
    #   金額は一致しているので結果は同じだが、明示しておく。
    kame_owned = {(t, r, m) for t, r, m, *_ in kame_rows}

    # ===== 人件費・法定福利費（給与データ 4月〜7月） =====
    F_PAY = PatternFill("solid", fgColor="E2D9F2")
    pay_cells = 0
    for tab, plrow, m, val, note in payroll.rows():
        if plrow not in build2.RIDX[tab]:
            missing.append((tab, plrow, "給与")); continue
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_PAY; c.number_format = build2.NUMFMT
        pay_cells += 1
    print("人件費セル", pay_cells)

    # ===== 売上（既存スプシから転記） =====
    F_SALES = PatternFill("solid", fgColor="FDE9D9")
    sales_cells = 0
    for tab, rows in sales.SALES.items():
        for rowname, vals in rows.items():
            r = build2.RIDX[tab][rowname]
            for i, m in enumerate(build2.MONTHS):
                if vals[i] is None:
                    continue
                if (tab, rowname, m) in board_owned:
                    continue          # boardを正とするので既存スプシ値は使わない
                if (tab, rowname, m) in kame_owned:
                    continue          # かめや精算書を正とする（焼きたて屋の売上・消費税）
                c = wb[tab][f"{build2.MCOL[m]}{r}"]
                c.value = int(vals[i]); c.fill = F_SALES; c.number_format = build2.NUMFMT
                sales_cells += 1
    print("売上セル", sales_cells)

    # ===== 既存21期PLからの穴埋め（★いちばん最後に置くこと）=================
    # 請求書・カード・銀行から積み上げても2割しか埋まらなかった。既存PLは
    # 1年ぶん手で入れてある記録なので、書類から作れないセルはこれを写す。
    # ★空いているセルだけ。書類から入れた値は絶対に上書きしない。
    #   だからこのブロックは全部の転記が終わったあとに置く必要がある。
    #   重なって金額が違うセルは exist_fill.conflicts() が拾い、下の
    #   「既存PLとの食い違い」タブに出す。判断は利用者に委ねる。
    ef_conf = exist_fill.conflicts(wb)
    # ★カード（F_CARD）と同じ色にしないこと。どこから来た値か見分けがつかなくなる
    F_EXIST = PatternFill("solid", fgColor="DDEBF7")
    ef_rows = list(exist_fill.rows(wb))
    for tab, plrow, m, val, src, note in ef_rows:
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + val; c.fill = F_EXIST; c.number_format = build2.NUMFMT
    # ★カード明細のルール（利用者指示 2026-08-21）を書き込みのあとに検算する。
    #   明細のある月は費目別。一括行に入っていたら二重計上なので止める。
    exist_fill.check(wb)
    ef_lump = exist_fill.lump_months(wb)
    print("既存PLからの穴埋めセル", len(ef_rows), "／計", f"{sum(x[3] for x in ef_rows):,}",
          "／食い違い", len(ef_conf), "セル",
          "／カード一括のまま", len(ef_lump), "セル（明細待ち）")

    # 明細ログ
    ws = wb.create_sheet("明細ログ")
    ws.cell(1, 1, "明細ログ（転記の元データ／検証用）").font = Font(bold=True, size=13)
    hdr = ["日付","店舗","取引先","摘要","税込","税率","税抜","消費税","転記先PL行","転記先の月",
           "データ元ファイル名","処理日時","備考"]
    for j, h in enumerate(hdr, start=1):
        c = ws.cell(2, j, h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496"); c.border = BORD
    for _, x in live.iterrows():
        ws.append([x["利用日"], x["店舗"], x["取引先"], x["摘要"], int(x["税込"]), int(x["税率"]),
                   int(x["税抜"]), int(x["消費税"]),
                   REMAP.get((x["店舗"], x["PL行"]), x["PL行"]), x["計上月"], x["元ファイル"], STAMP,
                   "" if pd.isna(x.get("備考")) else x.get("備考")])
    for _, x in ok[ok["判定"] == "除外"].iterrows():
        ws.append([x["利用日"], x["店舗"], x["取引先"], x["摘要"], int(x["税込"]), "", "", "",
                   "（除外）二重計上防止", "", x["元ファイル"], STAMP, x["備考"]])
    for tab, vendor, plrow, ex, tax, src, biko in INV:
        ws.append(["2026-07", tab, vendor, "請求書", ex + tax, "", ex, tax, plrow, "7月",
                   f"買掛/21期/2607月/{src}", STAMP, biko])
    for tab, vendor, plrow, ex, tax, src, biko in inv6.INV6:
        ws.append(["2026-06", tab, vendor, "請求書", ex + tax, "", ex, tax, plrow, "6月",
                   f"買掛/21期/2606月/確認済/{src}", STAMP, biko])
    for tab, vendor, plrow, ex, tax, src, biko in inv8.INV8:
        ws.append(["2026-08", tab, vendor, "請求書", ex + tax, "", ex, tax, plrow, "8月",
                   f"買掛/21期/{src}", STAMP, biko])
    for tab, plrow, m, val, src, note in nihonshokken.rows():
        ws.append([f"21期 {m}", tab, "日本食研", "請求書", "", 8, val, "", plrow, m,
                   src, STAMP, note])
    for tab, vendor, src, biko in JISSEKI:
        ws.append(["2026-07", tab, vendor, "実績", "", "", "", "", "（金額なし・実績のみ）", "7月",
                   f"買掛/21期/2607月/{src}", STAMP, biko])
    for tab, merch, plrow, ex, tax, src, m, used, iss in cards.rows():
        ws.append([used, tab, merch, f"{iss}カード", ex + tax, 10, ex, tax,
                   plrow or "（保留）", m, src, STAMP,
                   "支払日ベースで計上（明細は利用日をそのまま記載）。"
                   "カード明細に税額の記載がないため10%で逆算"])
    for tab, plrow, m, ex, tax, n, src, detail in board_rows:
        for d, cust, v in detail:
            ws.append([d, tab, cust, "board請求", "", 10, v, "", plrow, m, src, STAMP,
                       f"boardの請求一覧CSVより（グループ列で部門判定）。この行を含む{n}件を合計{ex:,}円として転記"])
    for tab, plrow, m, val, src, (fee, fee_tax) in demae_rows:
        ws.append(["", tab, "出前館", "支払通知書", fee, 10, val, fee_tax, plrow, m, src, STAMP,
                   "出前館利用料⑥の税抜（サービス利用料10%＋配達代行25%＋振込手数料＋決済手数料）。"
                   "既存PLと11か月とも一致することを確認済み（demaekan.EXIST_21_FEE）"])
    for tab, plrow, m, val, src, back in demae_refund:
        ws.append(["", tab, "出前館", "支払通知書", back, "", val, "", plrow, m, src, STAMP,
                   "お戻し金額⑦（商品代金補填・不課税）を費用のマイナスで計上。"
                   "既存PLも同じ扱い（年計 ▲20,890）"])
    for tab, plrow, m, val, src in kame_rows:
        ws.append(["", tab, "かめや", "本部請求", "", "", val, "", plrow, m, src, STAMP, ""])
    for tab, plrow, m, val, src, note in kame_cash:
        ws.append(["", tab, "かめや", "合計精算書", "", "", val, "", plrow, m, src, STAMP, note])
    for tab, plrow, m, val, src in yoko_rows:
        d = yokocho.DATA[m][tab]
        note = (f"神栖横丁が入居店舗へ出す請求書 No.{d['No']}（内訳: "
                f"{yokocho.detail(tab, plrow, m)}）。"
                f"請求書の小計{d['小計']:,}／税込{d['税込']:,}。"
                "神栖横丁側の売上は board.py が売掛CSVから入れている")
        if d.get("相殺の税抜補正"):
            note += "／★" + d["相殺の税抜補正"]
        ws.append(["", tab, "神栖横丁", "社内請求", "", 10, val, "", plrow, m, src,
                   STAMP, note])
    # ★損益に載せなかったスポンサー売掛相殺も記録として残す（転記先PL行は空欄）
    for tab, m, amt, src, item in yokocho.sponsor():
        # 転記先PL行に「その他経費」を書いておく。金額（税抜）は空のままなので、
        # cellnote.py が「※PLに載せていないもの」としてメモの末尾に出してくれる。
        ws.append(["", tab, "神栖横丁", "社内請求", "", 10, "", "", "その他経費", m, src, STAMP,
                   f"{item} {amt:,}（税抜）。★PLには載せていない（利用者確認 2026-08-23）。"
                   "スポンサーへの請求は横丁が出して入金も横丁に入り、"
                   "その分を各店舗への請求から差し引いている。店舗はレジを通していて"
                   "売上は当日計上済みなので、この相殺は売掛金と買掛金の相殺＝貸借だけの動き。"
                   "費用のマイナスにすると相殺の分だけ店舗の利益が多く出てしまう。"
                   "神栖横丁側は board のテナント請求が相殺後の小計なので、"
                   "スポンサー請求を売上に立てて合計がちょうど本来の家賃収入になる"])
    for tab, plrow, m, val, src, note in yb_rows:
        ws.append([f"21期 {m}", tab, note.split("。")[0], "口座振替", "", 10, int(val), "",
                   plrow, m, src, STAMP, note])
    for tab, plrow, m, val, src, note in sb_rows:
        ws.append([f"21期 {m}", tab, note.split("。")[0], "口座振替", "", "", int(val), "",
                   plrow, m, src, STAMP, note])
    for tab, plrow, m, val, src, note in tr_rows:
        ws.append([f"21期 {m}", tab, note.split("。")[0], "振込", "", "", int(val), "",
                   plrow, m, src, STAMP, note])
    for tab, plrow, m, val, src, note in nr_rows:
        ws.append([[p[5] for p in norow.POSTINGS
                    if (p[0], p[1], p[2]) == (tab, plrow, m)][0], tab,
                   "MEG design office" if "MEG" in plrow else "ヒラノ　タカシ"
                   if "お米" in plrow else plrow,
                   "請求書" if "MEG" in plrow else "銀行明細",
                   "", 10 if "MEG" in plrow else "", int(val), "",
                   plrow, m, src, STAMP, note])
    for tab, plrow, m, val, src, note in car_rows:
        ws.append([f"21期 {m}", tab, "陸事総合協同組合" if "陸自" in plrow else "トヨタファイナンス",
                   "請求書", "", "", int(val), "", plrow, m, src, STAMP, note])
    for tab, plrow, m, val, src, note in fix_rows:
        ws.append(["", tab, plrow, "自動引落（定額）", "", "", val, "", plrow, m, src, STAMP,
                   note or "既存スプシで11か月とも同額。請求書は来ない"])
    for tab, plrow, m, val, src, note in kaihi_rows:
        ws.append(["2026-08-16", tab, note.split(" — ")[0], "請求書", "", 10, val, "",
                   plrow, m, src, STAMP, note.split(" — ", 1)[1]])
    # ★既存21期PLからの穴埋めもログに出す（2026-08-23）。
    #   これを出さないと、セルのメモ（cellnote.py）が付かないセルが1,000件残る。
    for tab, plrow, m, val, src, note in ef_rows:
        ws.append(["", tab, "（既存21期PL）", "既存PL", "", "", val, "", plrow, m, src,
                   STAMP, note])
    # ★わざと写さなかったセルも記録に残す（2026-08-23）。金額（税抜）は空にするので、
    #   cellnote.py が「※PLに載せていないもの」としてメモの末尾に出す。
    #   空欄を見て「なぜ入っていないのか」を探さずに済むようにするため。
    for tab, plrow, m, v, why in exist_fill.not_posted_rows():
        ws.append(["", tab, "（既存21期PL）", "既存PL", "", "", "", "", plrow, m,
                   f"既存21期PL {tab}", STAMP, why])
    for tab, plrow, m, val, note in payroll.rows():
        ws.append([f"2026-{ {'4月':'04','5月':'05','6月':'06','7月':'07'}[m] }", tab, "（給与）",
                   "人件費" if plrow.startswith("人件費") else "社会保険料",
                   val, "", val, 0, plrow, m, payroll.SRC, STAMP, note])
    for col, w in zip("ABCDEFGHIJKLM", [11,14,26,12,10,6,10,9,26,10,42,18,60]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"

    # 保留リスト
    hs = wb.create_sheet("保留リスト")
    hs.cell(1, 1, "保留リスト（判定できなかったもの／推測で埋めていません）"
            ).font = Font(bold=True, size=13, color="C00000")
    for j, h in enumerate(["区分","日付/月","店舗","取引先","金額(税込)","理由・メモ"], start=1):
        c = hs.cell(2, j, h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="C00000"); c.border = BORD
    for _, x in hold.iterrows():
        hs.append(["カード明細", x["利用日"], x["店舗"], x["取引先"], int(x["税込"]), x["理由"]])
    for m, item, reason in kameya.hold_rows():
        hs.append(["かめや", m, "焼きたて屋", item, "", reason])
    for m, tab, item, reason in yokocho.hold_rows():
        hs.append(["横丁社内請求", m, tab, item, "", reason])
    for m, tab, item, reason in yokocho_bank.hold_rows():
        hs.append(["横丁口座", m, tab, item, "", reason])
    for m, tab, item, reason in store_bank.hold_rows():
        hs.append(["店舗口座", m, tab, item, "", reason])
    for m, tab, item, reason in namefa.hold_rows():
        hs.append(["なめがた", m, tab, item, "", reason])
    for m, tab, item, reason in shiina.hold_rows():
        hs.append(["産廃", m, tab, item, "", reason])
    for m, tab, item, reason in norow.hold_rows():
        hs.append(["新設行", m, tab, item, "", reason])
    # ★8月は書類だけで組む（利用者指示 2026-08-23「既存は無視してくださいね」）。
    #   まだ届いていない元データを保留リストに出して、何を待っているかを見えるようにする。
    for m, tab, item, reason in status8.hold_rows():
        hs.append(["8月待ち", m, tab, item, "", reason])
    for m, tab, item, reason in exist_fill.hold_rows(wb):
        hs.append(["既存PL", m, tab, item, "", reason])
    for m, tab, item, reason in transfers.hold_rows():
        hs.append(["振込", m, tab, item, "", reason])
    for m, tab, item, reason in rikuji.hold_rows():
        hs.append(["陸事総合", m, tab, item, "", reason])
    for tab, m, vendor, amount, reason in EXTRA_HOLD:
        hs.append(["請求書", m, tab, vendor, amount, reason])
    for iss, merch, ex, used in card_hold:
        hs.append([f"{iss}カード", used, "本部", merch, ex, "取引先マスタ（cards.py）に未登録。行名を要指示"])
    for tab, vendor, src, reason in INV_HOLD:
        hs.append(["請求書", "", tab, vendor, "", f"{src} — {reason}"])
    if hs.max_row == 2:
        hs.cell(3, 1, "保留はありません（2026-08-20 時点／カード明細6月・7月＋請求書2606月・2607月＋給与4〜7月）"
                ).font = Font(bold=True, color="008000")
    for col, w in zip("ABCDEF", [12,12,16,26,12,95]):
        hs.column_dimensions[col].width = w
    hs.freeze_panes = "A3"

    # ===== 既存PLとの食い違い（2026-08-21 新設）==============================
    # 新シートは書類（請求書・カード明細・銀行明細）から作った値を正としている。
    # 既存21期PLと金額が違うセルはここに全部出す。上書きはしていない。
    # 「差」がプラス＝新シートのほうが大きい。
    cs = wb.create_sheet("既存PLとの食い違い")
    cs.cell(1, 1, "既存21期PLと金額が違うセル（新シートは書類の値を採っている。"
                  "上書きはしていない）").font = Font(bold=True, size=12)
    for j, h in enumerate(["タブ", "行", "月", "新シート（書類）", "既存21期PL", "差",
                           "備考"], 1):
        c = cs.cell(2, j, h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="ED7D31")
    # ★利用者に判断してもらった食い違いは、そのむね書いておく。
    #   毎回「未解決の差」に見えてしまうと、本当に見るべきものが埋もれる。
    yoko_cells = {(t, p, m) for t, p, m, _v, _s in yokocho.rows()}
    for tab, plrow, m, new, old in sorted(ef_conf, key=lambda x: -abs(x[3] - x[4])):
        biko = ""
        if (tab, plrow, m) in yoko_cells:
            d = yokocho.DATA[m][tab]
            biko = (f"神栖横丁の請求書 No.{d['No']} を採用（利用者指示 2026-08-23"
                    f"「請求書で合わせて」）。内訳: {yokocho.detail(tab, plrow, m)}")
        cs.append([tab, plrow, m, new, old, new - old, biko])
    for col, w in zip("ABCDEFG", (14, 26, 6, 16, 16, 14, 90)):
        cs.column_dimensions[col].width = w
    for row in cs.iter_rows(min_row=3, min_col=4, max_col=6):
        for c in row:
            c.number_format = build2.NUMFMT
    cs.freeze_panes = "A3"

    # ===== セルのメモ（利用者提案 2026-08-23「コメント欄を利用してはどうかな」）=====
    # 明細ログを (店舗, PL行, 月) で束ねて、数字が入っているセルにメモを付ける。
    # 「この数字なんだっけ？」をセルの上で解決できるようにするため。
    # ★明細ログが唯一の元ネタ。束ねた合計がセルの値と合わなければ★印が出る。
    n_note, n_bad = cellnote.build(wb)
    print("セルのメモ", n_note, "件"
          + (f" ／★明細とセルの値が合わない {n_bad} 件" if n_bad else ""))

    wb.save(dst)
    print("転記セル数", posted, "／ 明細ログ", ws.max_row - 2, "行 ／ 保留", hs.max_row - 2, "件")
    if missing:
        print("!! 行が見つからない:", missing)


if __name__ == "__main__":
    main()
