#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""8月分（21期の最終月）の受け入れ状況をひと目で見るための一覧

    python3 status8.py

利用者方針「8月分は全部、CSVかPDFで入れていくのでそれ見て反映していきましょう」
（2026-08-21）。資料が1つ届くたびにこれを流して、何が入って何がまだかを見る。

21期は 2025年9月〜2026年8月。ほかの月は 9月〜7月 の11か月で組んであり、
8月だけが最後に残っている。モジュールによって「8月を拾えるか」が違うので、
ここに全部並べる。

--- 月のずれ（元データの年月とPL列の関係）-------------------------------
    千葉銀行/PayPay 明細   ファイルの年月＝取引月。支払日で1〜2か月戻す
    freeeカード statement  ファイル年月 −1 か月 ＝ 利用月（PL列）
    JCB/三井住友            ファイル年月 −1 か月（支払日ベース。cards.py 参照）
    陸事総合                ファイル年月＝利用月＝PL列
    ENEOS                  ファイル年月＝引落月。PL列は −2 か月
    なめがた請求書          件名の「N月度分」＝PL列
    買掛の請求書            フォルダ「26MM月」＝PL列
"""
import glob
import os

BASE = os.path.dirname(os.path.abspath(__file__))
M = "8月"

# ★2026-08-23 に Dropbox を実地で見て回った結果を DROPBOX に書いてある。
#   利用者指示「8月は基本、CSVやPDFから全部数字入れてください。既存は無視して
#   くださいね。足りないところは随時確認してください」。
#   既存21期PLは8月の値を1つも持っていないので、書類が来ないと8月は埋まらない。
#   exist_fill.NO_EXIST_MONTHS で8月は既存PLから写さないようにもしてある。

# (区分, 期待するファイル, 置き場所, 8月に要るもの, コードが8月を拾えるか, Dropboxの状況)
SOURCES = [
    ("銀行", "千葉銀行9口座 202608", "bank/小見川支店_普通_*_202608.csv",
     "小見川支店_普通_*_202608.csv", "自動（ファイルを置けば拾う）",
     "★横丁(3543920)の202608だけある。ほかの8口座は202607が最後"),
    ("銀行", "PayPay銀行", "bank/NBG_2026.csv",
     "NBG_2026.csv に8月の行", "自動（ファイルを置けば拾う）",
     "★Dropbox版は2026/07/08まで。手元の版(7/31まで)のほうが新しい"),
    ("カード", "freeeカード", "csv/statement-2026-09.csv",
     "statement-2026-09.csv（9月ファイル＝8月利用分）", "engine.py に1行足す",
     "★まだ。statement-2026-08.csv が最後"),
    ("カード", "JCB", "cards/202609meisai.csv",
     "202609meisai.csv（9月支払＝8月列）", "★対応済み",
     "あり（2026-08-25 到着）。64件・8月列へ。明細の和 421,284 が"
     "ファイルの「今回のお支払金額合計」と一致"),
    ("カード", "三井住友", "cards/202609.csv",
     "202609.csv（9月支払＝8月列）", "★対応済み",
     "あり（202609.csv）。本部の2件が8月に入っている"),
    ("車両", "陸事総合", "rikuji_pdf/202608.pdf",
     "0000042400_高速_請求書_202608.pdf", "rikuji.py の PDF_DATA に1行足す",
     "★まだ"),
    ("車両", "ENEOS", "eneos/ENEOS_8月.csv",
     "ENEOS_202610.csv（8月利用＝10月引落）", "eneos.py に1行足す",
     "★まだ。ENEOS_202608.csv（＝6月利用分）が最後"),
    ("請求書", "なめがたしろはとファーム", "namefa/8月.pdf",
     "なめがたしろはとふぁーむ　十三里屋9月末.pdf", "namefa.py の INVOICE に1行足す",
     "★まだ（買掛/21期/2608月/十三里屋/ が空）"),
    ("請求書", "買掛 2608月フォルダ一式", "（Dropbox）買掛/21期/2608月/",
     "各店舗フォルダのPDF", "読んで inv8.py に足す",
     "★3枚だけ（かりすまーず・JBQ・ピカピカ）。全部 inv8.py に入れ済み。"
     "十三里屋・焼きたて屋・もも焼き・タコハイ・ハナ・りゅうちゃん・"
     "鳥害対策・業務の各フォルダは空"),
    ("請求書", "神栖横丁→入居店舗の請求書", "（Dropbox）買掛/21期/2608月/<店舗>/",
     "横丁　<店舗>9月末.pdf（4店舗）", "yokocho_parse.py が読む",
     "★まだ。9月末請求なので発行前とみられる"),
    ("売掛", "board 請求書CSV", "cards/invoices*.csv",
     "8月ぶんを含む invoices.csv", "board.py の FILES に1行足す",
     "invoices.csv はあるが7月ぶん。8月の合計請求書Noが入ったものが要る"),
    ("その他", "かめや日報", "（Dropbox）", "8月ぶん", "kameya.py の MONTHS_21 に足す",
     "★まだ"),
    ("その他", "出前館", "（Dropbox）", "8月ぶん", "demaekan.py に足す", "★まだ"),
    # 給与のPDFは /※プロスタイル給与※/プロスタイル給与R8年/ に置かれる。
    # 8月分は9月上旬に出るはず（7月分は7/19作成・7/24支給だった）。
    # 8月分が届いたら kyuyo_parse.names() で氏名を取り、roster.py（勤怠アプリの
    # 名簿）で店舗に割る。202607で56人全員割り当てできることを確認済み。
    ("給与", "給料一覧表", "kyuyo/202608.pdf",
     "給料一覧表-202608.pdf（8月分）", "★対応済み（payroll8.py）",
     "あり（2026-08-24 到着）。56人・総支給 7,317,242／社会保険 618,404。"
     "名簿と部門コードで10タブに割り振り、22セルに計上"),
]


def hold_rows():
    """まだ来ていない8月の資料を保留リストへ出す。"""
    for kind, name, _where, want, _how, state in SOURCES:
        if not state.startswith("★"):
            continue
        yield ("8月", f"（{kind}）", name,
               f"8月は書類だけで組む方針（利用者指示 2026-08-23「既存は無視して"
               f"くださいね」）。{want} がまだ Dropbox に無いので、この元データ"
               f"ぶんの8月は空のまま。{state.lstrip('★')}")

# 定額（請求書が無く毎月同額）。8月も同額と決めてよいかは銀行明細で裏を取ってから。
FIXED_NOTE = ("fixed_costs.py は9〜7月の11か月で組んである。8月も同額のはずだが、"
              "銀行明細（202608）で裏を取ってから足すこと。推測で埋めない。")


def _has(pattern):
    if pattern.startswith("（"):
        return None                      # ローカルには来ないもの
    hit = glob.glob(os.path.join(BASE, pattern))
    return sorted(os.path.basename(h) for h in hit)


def main():
    import warnings
    warnings.filterwarnings("ignore")
    import openpyxl
    import build2
    print(f"■ {M}分の受け入れ状況\n")
    print(f"{'区分':<6}{'元データ':<24}{'手元':<6}{'Dropboxの状況（2026-08-23 実地確認）'}")
    print("-" * 116)
    for kind, name, pattern, need, how, state in SOURCES:
        got = _has(pattern)
        mark = "―" if got is None else ("あり" if got else "★まだ")
        print(f"{kind:<6}{name:<24}{mark:<6}{state}")
    print("-" * 116)
    miss = [x for x in SOURCES if x[5].startswith("★")]
    print(f"\n★まだ来ていない元データ {len(miss)}件 ／ 全{len(SOURCES)}件")
    print("要るファイル:")
    for kind, name, _p, want, how, _s in miss:
        print(f"   [{kind}] {name}: {want}")
        print(f"        → 届いたら {how}")
    print(f"\n定額: {FIXED_NOTE}")

    wb = openpyxl.load_workbook("損益計算書_21期テスト版.xlsx")
    print(f"\n■ いま{M}列に入っているもの")
    tot = 0
    for tab in build2.RIDX:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        hit = [(p, int(v)) for p, r in build2.RIDX[tab].items()
               if (v := ws[f"{build2.MCOL[M]}{r}"].value) and not isinstance(v, str)]
        if hit:
            print(f"   {tab}: " + " ／ ".join(f"{a} {b:,}" for a, b in hit))
            tot += sum(b for _, b in hit)
    print(f"   計 {tot:,}円" if tot else "   （まだ何も入っていない）")


if __name__ == "__main__":
    main()
