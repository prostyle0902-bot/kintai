#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""22期（2026.9〜2027.8）の空シートを作る

21期テスト版とまったく同じ11タブ・同じ行構成・同じ数式で、データは空。
2026年9月から本番運用するための器。

    python3 build22.py

行を増やしたくなったら build2.py の EXTRA_COGS / EXTRA_SGA を直す。
21期と22期で同じ定義を共有しているので、両方に反映される。

★2026-09-02 から、空の器ではなくなった。inv22.py に書いたぶんを入れる。
  21期の請求書フォルダに届いたものでも、費用の期間が22期にかかるものは
  21期に入れず inv22.py へ回す（第1号はクリーンシステム科学研究所の年間購読料）。
"""
from openpyxl.styles import PatternFill

import build2
import inv22

OUT = "損益計算書_22期.xlsx"
F_POST = PatternFill("solid", fgColor="FFF9C4")

if __name__ == "__main__":
    wb = build2.new_wb(period="22期")
    inv22.check(wb)
    n = 0
    for tab, plrow, m, ex, _tax, _v, _src, _b in inv22.rows():
        c = wb[tab][f"{build2.MCOL[m]}{build2.RIDX[tab][plrow]}"]
        c.value = int(c.value or 0) + ex
        c.fill = F_POST
        c.number_format = build2.NUMFMT
        n += 1
    wb.save(OUT)
    print(f"{OUT} を作成 ／ 転記 {n} セル ／ 計 "
          f"{sum(r[3] for r in inv22.rows()):,}円（inv22.py）")
    for t in build2.TABS:
        print(f"  {t:<14} {len(build2.LAYOUTS[t]):>3}行")
    print(f"  タブ: {wb.sheetnames}")
